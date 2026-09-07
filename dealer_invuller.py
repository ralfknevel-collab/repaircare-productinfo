"""
Kern van de dealer-Excel invuller (geen Streamlit-code).

Leest een dealerbestand (.xlsx/.csv), vindt de kopregel, matcht artikelen via de
mapping, vult lege cellen met productdata en voegt een tabblad 'Controle' toe.
Ook bruikbaar als script:

    python3 dealer_invuller.py dealerbestand.xlsx [--mapping mapping.json] [--overschrijven]
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import io
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from copy import deepcopy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import anthropic
import httpx
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

from artikeldata import Artikeldata, ComponentMaten, Match, Waarde, doosinhoud, normaliseer_code, normaliseer_ean
from mapping import KolomMapping, Mapping, lege_mapping, vraag_mapping
from veldcatalogus import catalogus_voor_prompt, converteer, veld

SLEUTELTYPE_ARG = {
    "sleutel_artikelcode": "artikelcode",
    "sleutel_ean": "ean",
    "sleutel_omschrijving": "omschrijving",
}

# Alleen eenduidige leveranciers- en EAN-koppen; een dealer-ArtNr is geen broncode.
SLEUTELKOPPEN = {
    "lieferantenartikelnummer": "sleutel_artikelcode",
    "liefmat": "sleutel_artikelcode",
    "herstellerartnr": "sleutel_artikelcode",
    "herstellerartikelnummer": "sleutel_artikelcode",
    "artikelnummerleverancier": "sleutel_artikelcode",
    "repaircareartikelnummer": "sleutel_artikelcode",
    "supplieritemnumber": "sleutel_artikelcode",
    "artikelcode": "sleutel_artikelcode",
    "ean": "sleutel_ean", "ean13": "sleutel_ean", "eancode": "sleutel_ean",
    "gtin": "sleutel_ean", "gtinean": "sleutel_ean",
}


def _norm_kop(waarde) -> str:
    return re.sub(r"[^a-z0-9]", "", str(waarde).casefold())


# Alleen volledige, bekende koppen. Een langer kenmerk zoals Farbechtheit is
# geen kleur; een doosgewicht is geen gewicht per stuk.
LOKALE_VELDKOPPEN = {
    _norm_kop(kop): doel
    for doel, aliassen in {
        "kleur": ("Kleur", "Farbe", "Farbton", "Colour", "Color"),
        "omschrijving": ("Omschrijving", "Productnaam", "Artikelnaam", "Artikelbezeichnung",
                         "Artikelname", "Artikelbezeichnung / Artikelname", "Produktname",
                         "ArtBeschreibung", "Product name", "Item name"),
        "gn_code": ("GN-code", "HS-code", "Douanetariefnummer", "Zolltarifnummer",
                    "Warentarifnummer", "Commodity code", "Tariff code", "Customs tariff number"),
        "netto_gewicht": ("Netto gewicht", "Nettogewicht", "Netto gewicht per stuk",
                          "Nettogewicht pro Stück", "Net weight", "Net weight per piece"),
        "bruto_gewicht": ("Bruto gewicht", "Bruttogewicht", "Bruto gewicht per stuk",
                          "Bruttogewicht pro Stück", "Gross weight", "Gross weight per piece"),
        "collo_netto_gewicht": ("Netto gewicht per doos", "Nettogewicht pro Karton", "Nettogewicht je Karton",
                                "Carton net weight", "Net weight per carton", "VPE Netto-Gewicht"),
        "collo_bruto_gewicht": ("Bruto gewicht per doos", "Bruttogewicht pro Karton", "Bruttogewicht je Karton",
                                "Carton gross weight", "Gross weight per carton", "VPE Brutto-Gewicht"),
        "lengte": ("Lengte", "Lengte per stuk", "Länge", "Laenge", "Länge pro Stück",
                   "Length", "Length per piece"),
        "breedte": ("Breedte", "Breedte per stuk", "Breite", "Breite pro Stück",
                    "Width", "Width per piece"),
        "hoogte": ("Hoogte", "Hoogte per stuk", "Höhe", "Hoehe", "Höhe pro Stück",
                   "Height", "Height per piece"),
        "collo_lengte": ("Collo lengte", "Lengte collo", "Kartonlänge", "Kartonlaenge", "Carton length"),
        "collo_breedte": ("Collo breedte", "Breedte collo", "Kartonbreite", "Carton width"),
        "collo_hoogte": ("Collo hoogte", "Hoogte collo", "Kartonhöhe", "Kartonhoehe", "Carton height"),
        "min_verkoophoeveelheid": ("Minimale verkoophoeveelheid", "Minimale afname",
                                  "Mindestabnahmemenge", "Minimum order quantity"),
        "adviesprijs": ("Adviesprijs excl. btw", "Verkoopadviesprijs excl. btw",
                        "UVP netto", "Recommended retail price excl. VAT"),
        "adviesprijs_eenheid": ("Adviesprijs eenheid",),
        "ve_aantal": ("VE volgens prijslijst",),
        "prijslijst_omschrijving": ("Omschrijving volgens prijslijst",),
    }.items()
    for kop in aliassen
}

LOKALE_EENHEDEN = {"g": "g", "gr": "g", "kg": "kg", "mm": "mm", "cm": "cm", "m": "m",
                   "stuk": "stuks", "stuks": "stuks", "stk": "stuks", "stück": "stuks",
                   "stueck": "stuks", "pcs": "stuks", "pieces": "stuks", "eur": "EUR", "€": "EUR"}

DOOSGEWICHT_VELDEN = {"collo_netto_gewicht", "collo_bruto_gewicht"}


def _vpe_gewichtkop(kop) -> bool:
    return bool(re.fullmatch(r"vpe(?:netto|brutto)gewicht(?:kg|g|gr)?", _norm_kop(kop)))


def _vpe_context(ws, kopregel: int, kolom: int):
    """Zoek eenduidige VPE-kolommen; een losse gewichtskop zegt niet of het om een doos gaat."""
    doel, kopeenheid, _ = _lokaal_doel(ws.cell(kopregel + 1, kolom).value)
    if doel not in DOOSGEWICHT_VELDEN:
        return None
    index = {}
    for cel in ws[kopregel + 1]:
        index.setdefault(_norm_kop(cel.value), []).append(cel.column)
    gewicht = "netto" if doel == "collo_netto_gewicht" else "brutto"
    aantallen, verpakkingen = index.get("vpeanzahl", []), index.get("vpeeinheit", [])
    eenheden = index.get(f"vpe{gewicht}gewichteinheit", [])
    if len(aantallen) != 1 or len(verpakkingen) != 1 or len(eenheden) > 1:
        return None
    if not eenheden and not kopeenheid:
        return None
    return aantallen[0], verpakkingen[0], eenheden[0] if eenheden else None, kopeenheid


def _gewichtseenheid(waarde):
    eenheid = LOKALE_EENHEDEN.get(str(waarde or "").strip().casefold().rstrip("."))
    return eenheid if eenheid in {"g", "kg"} else None


def _vpe_rijgegevens(ws, rij: int, context, artikel: dict, gekozen_eenheid):
    """Controleer de doosinhoud en gebruik uitsluitend de gewichtseenheid van deze rij."""
    if context is None:
        return None, "VPE-kolommen voor doosinhoud, verpakking of eenheid ontbreken of zijn dubbel."
    aantal_kolom, verpakking_kolom, eenheid_kolom, kopeenheid = context
    verpakking = str(ws.cell(rij, verpakking_kolom).value or "").strip().casefold()
    if verpakking not in {"karton", "doos", "carton", "box"}:
        return None, "VPE is geen eenduidige doos/karton; geen doosgewicht ingevuld."
    aantal = ws.cell(rij, aantal_kolom).value
    tekst = str(aantal).strip().replace(",", ".")
    if isinstance(aantal, bool) or not re.fullmatch(r"\d+(?:\.0+)?", tekst) or float(tekst) <= 0:
        return None, "Aantal stuks per VPE ontbreekt of is ongeldig."
    bron_aantal = doosinhoud(artikel)
    if bron_aantal is None:
        return None, "Doosinhoud ontbreekt in de productbron; minimale afname is geen doosinhoud."
    if float(tekst) != bron_aantal:
        return None, f"Dealer vraagt {tekst} stuks per doos; de bron beschrijft {bron_aantal}."
    eenheid = (_gewichtseenheid(ws.cell(rij, eenheid_kolom).value)
               if eenheid_kolom is not None else kopeenheid)
    if eenheid is None:
        return None, "Gewichtseenheid in de VPE-rij ontbreekt of is niet g/kg."
    if any(e is not None and e != eenheid for e in (kopeenheid, gekozen_eenheid)):
        return None, "Gekozen gewichtseenheid of kolomkop spreekt de eenheid in de VPE-rij tegen."
    return eenheid, f"Doosinhoud gecontroleerd: {bron_aantal} stuks; dealereenheid {eenheid}."


def _lokaal_doel(kop) -> tuple[str | None, str | None, str]:
    """Een veld en eenheid uit de kop zelf; voorbeeldgetallen zijn geen eenheidsbewijs."""
    tekst = " ".join(str(kop or "").split())
    doel = LOKALE_VELDKOPPEN.get(_norm_kop(tekst))
    # Het eurosymbool valt weg bij normaliseren; lees bij prijzen altijd de echte valuta.
    if doel == "adviesprijs":
        doel = None
    volledige_kop_herkend = doel is not None
    eenheid = None
    if doel is None:
        achtervoegsel = re.fullmatch(r"(.+?)\s*(?:\(([^()]+)\)|\[([^\[\]]+)\]|\s+(?:in\s+)?([^\s]+))",
                                     tekst, re.IGNORECASE)
        if achtervoegsel:
            doel = LOKALE_VELDKOPPEN.get(_norm_kop(achtervoegsel[1]))
            ruw = next(s for s in achtervoegsel.groups()[1:] if s is not None)
            eenheid = LOKALE_EENHEDEN.get(ruw.strip().casefold().rstrip("."))
            if doel and not veld(doel).eenheid:
                return None, None, "De toevoeging bij deze tekstkolom is niet eenduidig."
    v = veld(doel) if doel else None
    if v is None:
        return None, None, "Kolom niet eenduidig herkend; blijft leeg."
    if v.eenheid:
        if not eenheid:
            if volledige_kop_herkend and v.eenheid in {"g", "mm"}:
                return doel, None, "Herkend aan de volledige kolomkop; de kop vermeldt geen eenheid."
            return None, None, "Geen eenduidige ondersteunde eenheid in de kolomkop; blijft leeg."
        try:
            converteer(1, v.eenheid, eenheid)
        except ValueError:
            return None, None, "De eenheid in de kolomkop past niet bij dit veld; blijft leeg."
    return doel, eenheid, "Herkend aan de volledige kolomkop."


def _is_keuzeveld(ws, kolom: int, kopregel: int, data_start: int) -> bool:
    """PIM-keuzes en Excel-keuzelijsten vragen geen vrije producttekst."""
    for rij in range(1, kopregel + 1):
        if str(ws.cell(rij, kolom).value).strip().upper() in {"MW", "MM"}:
            return True
    for rij in range(kopregel + 2, data_start + 1):
        if str(ws.cell(rij, kolom).value).strip().upper().startswith("MWMEM"):
            return True
    for validatie in ws.data_validations.dataValidation:
        if validatie.type == "list" and any(
            bereik.min_col <= kolom <= bereik.max_col and bereik.max_row > data_start
            for bereik in validatie.sqref.ranges
        ):
            return True
    voorbeelden = {str(ws.cell(rij, kolom).value).strip().casefold()
                   for rij in range(data_start + 1, min(ws.max_row, data_start + 10) + 1)
                   if not _is_leeg(ws.cell(rij, kolom))}
    return bool(voorbeelden) and voorbeelden <= {"x", "✓", "✔"}


def ontbrekende_eenheden(mapping: Mapping) -> list[KolomMapping]:
    """Bekende gewicht- en maatvelden waarvoor nog geen eenheid gekozen is."""
    return [k for k in mapping.doelen()
            if k.eenheid is None and (v := veld(k.doelveld)) is not None and v.eenheid in {"g", "mm"}
            and not (k.doelveld in DOOSGEWICHT_VELDEN and _vpe_gewichtkop(k.kolom))]


def pas_eenheden_toe(mapping: Mapping, maat_eenheid: str | None = None,
                     gewicht_eenheid: str | None = None, bron: str = "Keuze gebruiker") -> Mapping:
    """Pas een bevestigde keuze toe op een kopie; expliciete kolomeenheden gaan voor."""
    if maat_eenheid is not None and maat_eenheid not in {"mm", "cm", "m"}:
        raise ValueError("Ongeldige maateenheid; kies mm, cm of m.")
    if gewicht_eenheid is not None and gewicht_eenheid not in {"g", "kg"}:
        raise ValueError("Ongeldige gewichtseenheid; kies g of kg.")
    gekozen = deepcopy(mapping)
    for k in ontbrekende_eenheden(gekozen):
        eenheid = maat_eenheid if veld(k.doelveld).eenheid == "mm" else gewicht_eenheid
        if eenheid is not None:
            k.eenheid = eenheid
            k.toelichting = (k.toelichting + f" Eenheid ingesteld op {eenheid} ({bron}).").strip()
    return gekozen


class _PUNTKOMMA(csv.excel):
    """Fallback-dialect als csv.Sniffer het scheidingsteken niet herkent."""
    delimiter = ";"


GEEL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
CONTROLE_TAB = "Controle"


def laad_werkboek(inhoud: bytes, bestandsnaam: str) -> openpyxl.Workbook:
    ext = Path(bestandsnaam).suffix.lower()
    if ext == ".xlsx":
        return openpyxl.load_workbook(io.BytesIO(inhoud))
    if ext == ".csv":
        try:
            tekst = inhoud.decode("utf-8-sig")
        except UnicodeDecodeError:
            tekst = inhoud.decode("cp1252")   # veelgebruikte export van Duitse ERP-systemen
        try:
            dialect = csv.Sniffer().sniff(tekst[:2000], delimiters=";,\t")
        except csv.Error:
            dialect = _PUNTKOMMA
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for rij in csv.reader(io.StringIO(tekst), dialect):
            ws.append([c if c != "" else None for c in rij])
        return wb
    raise ValueError(f"Bestandsformaat {ext or '(geen)'} wordt niet ondersteund. Sla het bestand op als .xlsx of .csv.")


def _heeft_data(ws) -> bool:
    return any(any(c is not None for c in rij) for rij in ws.iter_rows(values_only=True))


def _is_formuleblad(rijen: list[list]) -> bool:
    """Meer dan de helft van de gevulde cellen is een formule: hulpblad, geen invulblad."""
    gevuld = [c for rij in rijen for c in rij if c is not None and str(c).strip() != ""]
    formules = [c for c in gevuld if isinstance(c, str) and c.startswith("=")]
    return bool(gevuld) and len(formules) * 2 > len(gevuld)


def _heeft_kopregel(rijen: list[list]) -> bool:
    try:
        vind_kopregel(rijen)
        return True
    except ValueError:
        return False


def kies_tabblad(wb: openpyxl.Workbook, naam: str | None):
    """Gegeven naam, anders het eerste tabblad dat op een invulblad lijkt.

    Voorkeur: een blad met een herkenbare kopregel en zonder formules in de
    eerste rijen — meer dan de helft van de cellen (hulpbladen zoals 'Attribuutlijst' trekken data via formules
    en zijn niet het blad dat de dealer ingevuld wil hebben). Daarna: eerste
    blad met een kopregel, dan eerste blad met data, dan het eerste blad.
    """
    if naam:
        return wb[naam]
    met_data = [ws for ws in wb.worksheets if _heeft_data(ws)]
    if not met_data:
        return wb.worksheets[0]
    eerste_rijen = {ws.title: lees_rijen(ws, 30) for ws in met_data}
    for ws in met_data:
        rijen = eerste_rijen[ws.title]
        if _heeft_kopregel(rijen) and not _is_formuleblad(rijen):
            return ws
    for ws in met_data:
        if _heeft_kopregel(eerste_rijen[ws.title]):
            return ws
    return met_data[0]


def lees_rijen(ws, n: int = 10) -> list[list]:
    uit = []
    for i, rij in enumerate(ws.iter_rows(values_only=True)):
        if i >= n:
            break
        uit.append(list(rij))
    return uit


def vind_kopregel(rijen: list[list]) -> int:
    # Leverancierscode en EAN geven meer houvast dan groepskoppen zoals EM/MW.
    kandidaten = []
    for i, rij in enumerate(rijen):
        teksten = [c for c in rij if isinstance(c, str) and c.strip() and not c.startswith("=")]
        score = sum(_norm_kop(c) in SLEUTELKOPPEN for c in teksten)
        if score and len(teksten) >= 2:
            kandidaten.append((score, -i))
    if kandidaten:
        return -max(kandidaten)[1]
    for i, rij in enumerate(rijen):
        teksten = [c for c in rij if isinstance(c, str) and c.strip() and not c.startswith("=")]
        if len(teksten) >= 3:
            return i
    raise ValueError("Geen kopregel gevonden in de eerste rijen (verwacht een rij met minstens 3 tekstkoppen).")


def bepaal_data_start(ws, kopregel_index: int) -> int:
    """Sla lege rijen en technische PIM-veldcodes onder de koppen over (0-based)."""
    for rij in ws.iter_rows(min_row=kopregel_index + 2):
        gevuld = [str(c.value).strip() for c in rij if not _is_leeg(c)]
        if not gevuld:
            continue
        if all(re.fullmatch(r"(?:MATNR|(?:MW)?MEM[BD]\d+(?:-\d+)?)", v) for v in gevuld):
            continue
        return rij[0].row - 1
    return ws.max_row


def koppen(ws, kopregel_index: int) -> dict[str, int]:
    rij = next(ws.iter_rows(min_row=kopregel_index + 1, max_row=kopregel_index + 1, values_only=True))
    uit: dict[str, int] = {}
    for i, c in enumerate(rij):
        naam = str(c).strip() if c is not None and str(c).strip() else f"Kolom {get_column_letter(i + 1)}"
        basis, n = naam, 2
        while naam in uit:
            naam = f"{basis} ({n})"
            n += 1
        uit[naam] = i
    return uit


@dataclass
class VeldResultaat:
    kolom: str
    veld_id: str
    waarde: object
    eenheid: str | None
    bron: str
    regel: str | None
    status: str  # ingevuld | leeg | bestaand | controleer | overgeslagen | onzeker | eenheid_nodig


@dataclass
class RijResultaat:
    rij: int                     # 1-based Excel-rijnummer
    sleutel: str
    match: Match | None
    velden: list[VeldResultaat] = field(default_factory=list)
    toelichting: str | None = None


@dataclass
class Rapport:
    rijen: list[RijResultaat]
    overgeslagen_kolommen: list[str] = field(default_factory=list)

    def samenvatting(self) -> dict:
        via: dict[str, int] = {}
        gaten_per_kolom: dict[str, int] = {}
        ingevuld = gaten = onzeker = eenheid_nodig = 0
        for r in self.rijen:
            if r.match:
                via[r.match.via] = via.get(r.match.via, 0) + 1
            for v in r.velden:
                if v.status == "ingevuld":
                    ingevuld += 1
                elif v.status == "onzeker":
                    onzeker += 1
                elif v.status == "eenheid_nodig":
                    eenheid_nodig += 1
                elif v.status == "leeg":
                    gaten += 1
                    gaten_per_kolom[v.kolom] = gaten_per_kolom.get(v.kolom, 0) + 1
        gevonden = sum(1 for r in self.rijen if r.match)
        return {
            "totaal": len(self.rijen), "gevonden": gevonden, "niet_gevonden": len(self.rijen) - gevonden,
            "via": via, "ingevuld": ingevuld, "gaten": gaten, "gaten_per_kolom": gaten_per_kolom,
            "onzeker": onzeker, "eenheid_nodig": eenheid_nodig,
        }


def maak_waarde(w: Waarde, eenheid_doel: str | None):
    """Bronwaarde omrekenen naar de gevraagde eenheid; getallen netjes afronden."""
    if isinstance(w.waarde, ComponentMaten):
        return " / ".join(
            f"{naam}: {maak_waarde(Waarde(getal, w.eenheid, w.bron), eenheid_doel)}"
            for naam, getal in w.waarde.waarden
        )
    if isinstance(w.waarde, bool) or not isinstance(w.waarde, (int, float)):
        return w.waarde
    getal = float(w.waarde)
    if w.eenheid and eenheid_doel:
        getal = converteer(getal, w.eenheid, eenheid_doel)
    getal = round(getal, 3)
    return int(getal) if getal.is_integer() else getal


def _is_leeg(cel) -> bool:
    return cel.value is None or (isinstance(cel.value, str) and not cel.value.strip())


def _verbreed_maatkolom(ws, cel) -> None:
    """Geef nieuwe A/B-tekst ruimte zonder buurkolommen of bestaande breedtes te wijzigen."""
    # Een gedeelde instelling laten staan, ook als de doelkolom midden in het bereik ligt.
    if any(d.min is not None and d.max is not None and d.min < d.max
           and d.min <= cel.column <= d.max for d in ws.column_dimensions.values()):
        return
    bestaand = ws.column_dimensions.get(cel.column_letter)
    breedte = (bestaand.width if bestaand is not None
               else ws.sheet_format.defaultColWidth or ws.sheet_format.baseColWidth)
    nodig = len(cel.value) + 2
    if nodig > breedte:
        ws.column_dimensions[cel.column_letter].width = nodig


def _datarijen(ws, mapping: Mapping):
    start = mapping.data_start_index
    if start is None:
        start = mapping.kopregel_index + 1
    if not 0 <= mapping.kopregel_index < ws.max_row or not mapping.kopregel_index < start <= ws.max_row:
        raise ValueError("Controleer de kopregel en de eerste artikelrij: de eerste artikelrij moet na de kopregel liggen.")
    for rijnr in range(start + 1, ws.max_row + 1):
        cellen = ws[rijnr]
        if any(not _is_leeg(c) for c in cellen):
            yield rijnr, cellen


def _zoek_match(cellen, mapping: Mapping, kolomindex: dict[str, int], artikeldata: Artikeldata):
    argumenten: dict[str, object] = {}
    delen = []
    conflicten: set[str] = set()
    for k in mapping.sleutels():
        i = kolomindex.get(k.kolom)
        if i is None:
            continue
        waarde = cellen[i].value
        delen.append("" if waarde is None else str(waarde))
        soort = SLEUTELTYPE_ARG[k.doelveld]
        if soort == "artikelcode":
            waarde = normaliseer_code(waarde)
        elif soort == "ean":
            waarde = normaliseer_ean(waarde)
        else:
            waarde = " ".join(str(waarde or "").split()).casefold()
        if not waarde:
            continue
        if soort in argumenten and argumenten[soort] != waarde:
            conflicten.add(soort)
        else:
            argumenten[soort] = waarde
    sleutel = " / ".join(delen)
    if conflicten:
        return None, sleutel, "Tegenstrijdige waarden in sleutelkolommen: " + ", ".join(sorted(conflicten))
    match = artikeldata.zoek(**argumenten)
    bronconflicten = match.artikel.get("bron_conflicten", []) if match else []
    return match, sleutel, " ".join(bronconflicten) or None


def match_rijen(ws, mapping: Mapping, artikeldata: Artikeldata) -> list[RijResultaat]:
    if not mapping.sleutels():
        raise ValueError("Geen sleutelkolom gekozen (artikelnummer, EAN of omschrijving).")
    kolomindex = koppen(ws, mapping.kopregel_index)
    uit = []
    sleutelindices = [kolomindex[k.kolom] for k in mapping.sleutels() if k.kolom in kolomindex]
    for rijnr, cellen in _datarijen(ws, mapping):
        if not any(not _is_leeg(cellen[i]) for i in sleutelindices):
            continue
        match, sleutel, toelichting = _zoek_match(cellen, mapping, kolomindex, artikeldata)
        uit.append(RijResultaat(rijnr, sleutel, match, toelichting=toelichting))
    return uit


def overgeslagen_kolommen(mapping: Mapping, kolomindex: dict[str, int]) -> list[str]:
    """Kolommen uit de mapping die niet in de kopregel staan (worden niet gevuld of gezocht)."""
    uit: list[str] = []
    for k in mapping.kolommen:
        if k.doelveld != "geen" and k.kolom not in kolomindex and k.kolom not in uit:
            uit.append(k.kolom)
    return uit


def _zwarte_cellen_overslaan(wb) -> bool:
    """Respecteer de expliciete legenda van het kenmerkenbestand."""
    if "Legende" not in wb.sheetnames:
        return False
    for rij in wb["Legende"].iter_rows(values_only=True):
        tekst = " ".join(str(c) for c in rij if c is not None).casefold()
        if "schwarzen zellen" in tekst and "nicht befüllt" in tekst:
            return True
    return False


def _zwarte_themas(wb) -> set[int]:
    if not wb.loaded_theme:
        return set()
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    kleuren = ET.fromstring(wb.loaded_theme).find("a:themeElements/a:clrScheme", ns)
    if kleuren is None:
        return set()
    # Celthema's gebruiken licht/donker; in het XML staat donker/licht eerst.
    volgorde = ["lt1", "dk1", "lt2", "dk2"] + [f"accent{i}" for i in range(1, 7)] + ["hlink", "folHlink"]
    zwart = set()
    for i, naam in enumerate(volgorde):
        element = kleuren.find(f"a:{naam}", ns)
        if element is not None and len(element):
            kleur = element[0].get("lastClr") or element[0].get("val", "")
            if kleur.upper() == "000000":
                zwart.add(i)
    return zwart


def _is_zwart(cel, zwarte_themas: set[int]) -> bool:
    if cel.fill.patternType != "solid":
        return False
    kleur = cel.fill.fgColor
    if kleur.tint != 0:
        return False
    if kleur.type == "theme":
        return kleur.theme in zwarte_themas
    if kleur.type == "rgb":
        return kleur.rgb[-6:].upper() == "000000"
    if kleur.type == "indexed" and 0 <= kleur.indexed < len(COLOR_INDEX):
        return COLOR_INDEX[kleur.indexed][-6:] == "000000"
    return False


def vul_in(ws, mapping: Mapping, artikeldata: Artikeldata, overschrijven: bool = False,
           behoud_sjabloon: bool = False) -> Rapport:
    rijen = match_rijen(ws, mapping, artikeldata)
    kolomindex = koppen(ws, mapping.kopregel_index)
    overgeslagen = overgeslagen_kolommen(mapping, kolomindex)
    doelen = [(k, kolomindex[k.kolom]) for k in mapping.doelen() if k.kolom in kolomindex]
    vpe_contexten = {i: _vpe_context(ws, mapping.kopregel_index, i + 1) for k, i in doelen
                    if k.doelveld in DOOSGEWICHT_VELDEN
                    and _vpe_gewichtkop(ws.cell(mapping.kopregel_index + 1, i + 1).value)}
    zonder_eenheid = {k.kolom for k in ontbrekende_eenheden(mapping)}
    zwart_overslaan = _zwarte_cellen_overslaan(ws.parent)
    zwarte_themas = _zwarte_themas(ws.parent) if zwart_overslaan else set()
    for r in rijen:
        for k, i in doelen:
            cel = ws.cell(row=r.rij, column=i + 1)
            reden = None
            if isinstance(cel, MergedCell):
                reden = "onderdeel van een samengevoegde cel"
            elif cel.data_type == "f":
                reden = "bestaande Excel-formule"
            elif zwart_overslaan and _is_zwart(cel, zwarte_themas):
                reden = "zwarte cel: volgens de legenda niet invullen"
            elif behoud_sjabloon and r.match and r.match.via == "omschrijving":
                reden = "artikel alleen via omschrijving gevonden; niet automatisch invullen"
            if reden:
                r.velden.append(VeldResultaat(k.kolom, k.doelveld, cel.value, k.eenheid,
                                              "dealer", reden, "overgeslagen"))
                continue
            w = artikeldata.waarde(r.match.artikel, k.doelveld) if r.match else None
            if not _is_leeg(cel) and not overschrijven:
                r.velden.append(VeldResultaat(k.kolom, k.doelveld, cel.value, k.eenheid,
                                              "dealer (bestaande waarde)", None, "bestaand"))
                continue
            if r.toelichting:
                r.velden.append(VeldResultaat(k.kolom, k.doelveld, None, k.eenheid,
                                              "dealer (sleutelkolommen)", r.toelichting, "onzeker"))
                continue
            eenheid, verpakkingsregel = k.eenheid, None
            if i in vpe_contexten and r.match:
                eenheid, verpakkingsregel = _vpe_rijgegevens(
                    ws, r.rij, vpe_contexten[i], r.match.artikel, k.eenheid,
                )
                if eenheid is None:
                    r.velden.append(VeldResultaat(k.kolom, k.doelveld, None, k.eenheid,
                                                  "dealer en productbron (doosinhoud/eenheid)",
                                                  verpakkingsregel, "onzeker"))
                    continue
            if behoud_sjabloon and k.kolom in zonder_eenheid:
                r.velden.append(VeldResultaat(k.kolom, k.doelveld, None, None,
                                              w.bron if w else "geen productwaarde beschikbaar",
                                              "Eenheid ontbreekt; kies eerst de maat- of gewichtseenheid.",
                                              "eenheid_nodig"))
                continue
            if w is not None and not w.eenduidig and (behoud_sjabloon or w.waarde is None):
                reden = w.onzeker_reden or "Productwaarde is niet eenduidig."
                if w.regel:
                    reden += " " + w.regel
                r.velden.append(VeldResultaat(k.kolom, k.doelveld, None, k.eenheid, w.bron, reden, "onzeker"))
                continue
            if w is None or w.waarde is None or w.waarde == "":
                if not behoud_sjabloon:
                    cel.fill = GEEL
                bron = "artikel niet gevonden" if r.match is None else "geen waarde in productdata"
                r.velden.append(VeldResultaat(k.kolom, k.doelveld, None, k.eenheid, bron, None, "leeg"))
                continue
            cel.value = maak_waarde(w, eenheid)
            if isinstance(w.waarde, ComponentMaten):
                _verbreed_maatkolom(ws, cel)
            status = "controleer" if r.match.via == "omschrijving" else "ingevuld"
            regel = " ".join(deel for deel in (w.regel, k.toelichting, verpakkingsregel) if deel) or None
            r.velden.append(VeldResultaat(k.kolom, k.doelveld, cel.value, eenheid, w.bron, regel, status))
    return Rapport(rijen, overgeslagen)


def schrijf_controle(wb, rapport: Rapport, behoud_sjabloon: bool = False) -> None:
    naam = CONTROLE_TAB
    if behoud_sjabloon:
        nummer = 2
        while naam.casefold() in {b.casefold() for b in wb.sheetnames}:
            naam = f"{CONTROLE_TAB} ({nummer})"
            nummer += 1
    elif CONTROLE_TAB in wb.sheetnames:
        del wb[CONTROLE_TAB]
    ct = wb.create_sheet(naam)
    s = rapport.samenvatting()
    ct.append(["Samenvatting"])
    ct.append([f"Rijen: {s['totaal']}", f"Gevonden: {s['gevonden']}", f"Niet gevonden: {s['niet_gevonden']}",
               f"Ingevuld: {s['ingevuld']}", f"Gaten: {s['gaten']}", f"Onzeker: {s['onzeker']}",
               f"Eenheid nodig: {s['eenheid_nodig']}"])
    ct.append(["Gevonden via: " + ", ".join(f"{k} {v}" for k, v in s["via"].items())])
    ct.append(["Gaten per kolom: " + ", ".join(f"{k} {v}" for k, v in s["gaten_per_kolom"].items())])
    if rapport.overgeslagen_kolommen:
        ct.append(["Overgeslagen kolommen (niet in kopregel): " + ", ".join(rapport.overgeslagen_kolommen)])
    ct.append([])
    ct.append(["Rij", "Sleutel", "Artikelcode", "Gevonden via", "Kolom", "Doelveld", "Waarde", "Eenheid",
               "Status", "Bron", "Rekenregel"])
    for r in rapport.rijen:
        if r.match is None:
            ct.append([r.rij, r.sleutel, None, "niet gevonden", None, None, None, None,
                       "onzeker" if r.toelichting else None, None, r.toelichting])
            continue
        code = r.match.artikel.get("artikelcode")
        via = r.match.via if r.match.score >= 1.0 else f"{r.match.via} ({r.match.score:.2f})"
        for v in r.velden:
            ct.append([r.rij, r.sleutel, code, via, v.kolom, v.veld_id, v.waarde, v.eenheid,
                       v.status, v.bron, v.regel])
    for kol, breedte in zip("ABCDEFGHIJK", (6, 26, 12, 14, 20, 22, 16, 8, 11, 34, 50)):
        ct.column_dimensions[kol].width = breedte


def werkboek_naar_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def controleer_eenheden(mapping: Mapping) -> list[str]:
    """Meldingen voor doelkolommen waarvan de gekozen eenheid niet bij het doelveld past.

    Het schema staat elke eenheid bij elk veld toe, dus 'cm' bij een gewicht komt voor.
    Zonder deze controle klapt het invullen halverwege op een ValueError uit converteer().
    """
    meldingen: list[str] = []
    for k in mapping.doelen():
        v = veld(k.doelveld)
        if v is None or not k.eenheid or not v.eenheid:
            continue
        try:
            converteer(1.0, v.eenheid, k.eenheid)
        except ValueError:
            meldingen.append(f"Kolom {k.kolom!r}: eenheid {k.eenheid} past niet bij {v.label} ({v.eenheid})")
    return meldingen


def verwerk(inhoud: bytes, bestandsnaam: str, mapping: Mapping, artikeldata: Artikeldata,
            tabblad: str | None = None, overschrijven: bool = False,
            behoud_sjabloon: bool = False) -> tuple[bytes, Rapport]:
    wb = laad_werkboek(inhoud, bestandsnaam)
    ws = kies_tabblad(wb, tabblad)
    rapport = vul_in(ws, mapping, artikeldata, overschrijven, behoud_sjabloon)
    schrijf_controle(wb, rapport, behoud_sjabloon)
    return werkboek_naar_bytes(wb), rapport


def bepaal_mapping(client, ws, artikeldata: Artikeldata,
                   voortgang: Callable[[str], None] | None = None,
                   ai_fouten_doorgeven: bool = False) -> Mapping:
    """Koppel bekende koppen lokaal; optioneel via Claude met een lokale terugval."""
    rijen = lees_rijen(ws, 30)
    try:
        kopregel = vind_kopregel(rijen)
    except ValueError as e:
        return Mapping(0, [], opmerkingen=str(e))
    kolomindex = koppen(ws, kopregel)
    namen = list(kolomindex)
    def lokale_mapping(opmerking: str) -> Mapping:
        m = lege_mapping(kopregel, namen)
        m.data_start_index = bepaal_data_start(ws, kopregel)
        for k in m.kolommen:
            kolom = kolomindex[k.kolom] + 1
            # De getoonde naam kan '(2)' bevatten om dubbele kolommen te onderscheiden.
            ruw = ws.cell(kopregel + 1, kolom).value
            doel = SLEUTELKOPPEN.get(_norm_kop(ruw))
            if doel:
                k.doelveld, k.zekerheid = doel, "hoog"
                k.toelichting = "Herkend aan de kolomkop."
            elif _is_keuzeveld(ws, kolom, kopregel, m.data_start_index):
                k.toelichting = "Keuzeveld of x-markering; geen vrije producttekst invullen."
            else:
                doel, eenheid, k.toelichting = _lokaal_doel(ruw)
                if doel in DOOSGEWICHT_VELDEN and _vpe_gewichtkop(ruw):
                    context = _vpe_context(ws, kopregel, kolom)
                    if context is None:
                        k.toelichting = "VPE-gewicht zonder eenduidige doosinhoud, verpakking en eenheidskolom."
                        continue
                    if eenheid is None:
                        eenheden = {_gewichtseenheid(ws.cell(rij, context[2]).value)
                                    for rij in range(m.data_start_index + 1, ws.max_row + 1)}
                        eenheid = next(iter(eenheden)) if len(eenheden) == 1 else None
                    k.toelichting = "VPE-doosgewicht; doosinhoud en gewichtseenheid worden per artikelrij gecontroleerd."
                if doel:
                    k.doelveld, k.eenheid, k.zekerheid = doel, eenheid, "hoog"
        m.opmerkingen = opmerking
        return m
    if client is None:
        return lokale_mapping("Bekende kolommen lokaal herkend. Onzekere velden blijven leeg; "
                              "je kunt ze handmatig koppelen.")
    catalogus = catalogus_voor_prompt(artikeldata.ruwe_kolommen, artikeldata.vaste_sleutels)
    try:
        m = vraag_mapping(client, rijen[:max(10, kopregel + 6)], ws.title, ws.max_row, catalogus, voortgang)
    except (asyncio.TimeoutError, anthropic.APITimeoutError, httpx.TimeoutException):
        if ai_fouten_doorgeven:
            raise
        return lokale_mapping("De maximale wachttijd voor automatische kolomherkenning is verstreken. "
                              "Bekende kolommen lokaal herkend. Kies de overige velden handmatig.")
    except (anthropic.APIError, httpx.TransportError, ValueError, KeyError) as e:
        if ai_fouten_doorgeven:
            raise
        return lokale_mapping(f"Automatische kolomkoppeling mislukt ({e}). Kies de velden handmatig.")
    if m.data_start_index is None:
        m.data_start_index = bepaal_data_start(ws, m.kopregel_index)
    # Claude bouwt kolomnamen uit het ruwe fragment: die kunnen in hoofdletters of
    # whitespace afwijken van koppen(), dat lege koppen ook nog een naam geeft.
    # Zonder deze reconciliatie valt zo'n kolom stilzwijgend buiten het invullen.
    namen = list(koppen(ws, m.kopregel_index).keys())
    losjes = {" ".join(n.split()).casefold(): n for n in namen}
    ruwe_koppen = next(ws.iter_rows(min_row=m.kopregel_index + 1,
                                   max_row=m.kopregel_index + 1, values_only=True))
    voorkomens: dict[str, list[str]] = {}
    for naam, ruw in zip(namen, ruwe_koppen):
        origineel = str(ruw).strip() if ruw is not None else naam
        voorkomens.setdefault(" ".join(origineel.split()).casefold(), []).append(naam)
    gebruikt: set[str] = set()
    gereconcilieerd, ontbreekt = [], []
    for k in m.kolommen:
        norm = " ".join(str(k.kolom).split()).casefold()
        # Exacte unieke namen gaan voor; herhaalde ruwe koppen krijgen elk hun eigen kolom.
        exact = k.kolom if k.kolom in namen else losjes.get(norm)
        echt = exact if exact is not None and exact not in gebruikt else next(
            (naam for naam in voorkomens.get(norm, []) if naam not in gebruikt), None,
        )
        if echt is None:
            ontbreekt.append(k.kolom)
            continue
        k.kolom = echt
        gebruikt.add(echt)
        gereconcilieerd.append(k)
    m.kolommen = gereconcilieerd
    if ontbreekt:
        melding = " ".join(f"Kolom '{n}' uit het Claude-voorstel niet gevonden in de kopregel."
                           for n in ontbreekt)
        m.opmerkingen = (m.opmerkingen + " " + melding).strip()
    # Kolommen die Claude niet noemde toevoegen als 'geen', zodat de UI compleet is.
    genoemd = {k.kolom for k in m.kolommen}
    for naam in namen:
        if naam not in genoemd:
            m.kolommen.append(lege_mapping(m.kopregel_index, [naam]).kolommen[0])
    return m


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Vul een dealer-Excelbestand met Repair Care-productdata.")
    p.add_argument("bestand")
    p.add_argument("--mapping", help="mapping.json gebruiken in plaats van Claude")
    p.add_argument("--schrijf-mapping", help="gebruikte mapping opslaan als JSON")
    p.add_argument("--overschrijven", action="store_true")
    p.add_argument("--tabblad")
    p.add_argument("--uit")
    args = p.parse_args(argv)

    pad = Path(args.bestand)
    inhoud = pad.read_bytes()
    artikeldata = Artikeldata.laad()
    wb = laad_werkboek(inhoud, pad.name)
    ws = kies_tabblad(wb, args.tabblad)

    if args.mapping:
        mapping = Mapping.uit_dict(json.loads(Path(args.mapping).read_text(encoding="utf-8")))
    else:
        client = anthropic.AsyncAnthropic() if os.environ.get("ANTHROPIC_API_KEY") else None
        mapping = bepaal_mapping(client, ws, artikeldata)
        if mapping.opmerkingen:
            print("Opmerking:", mapping.opmerkingen)
    if args.schrijf_mapping:
        Path(args.schrijf_mapping).write_text(json.dumps(mapping.naar_dict(), ensure_ascii=False, indent=1),
                                              encoding="utf-8")
    for k in mapping.kolommen:
        print(f"  {k.kolom:30} -> {k.doelveld:24} {k.eenheid or '':4} [{k.zekerheid}] {k.toelichting}")

    meldingen = controleer_eenheden(mapping)
    if meldingen:
        for melding in meldingen:
            print("Eenheid past niet:", melding)
        print("Pas de eenheid in de mapping aan en probeer opnieuw.")
        return 1

    try:
        uit_bytes, rapport = verwerk(inhoud, pad.name, mapping, artikeldata, ws.title, args.overschrijven)
    except ValueError as e:
        print(f"Invullen niet mogelijk: {e}")
        print("Tip: geef een mapping mee met --mapping, of kies een sleutelkolom.")
        return 1
    uit = Path(args.uit) if args.uit else pad.with_name(pad.stem + "_ingevuld.xlsx")
    uit.write_bytes(uit_bytes)
    s = rapport.samenvatting()
    print(f"Geschreven: {uit}")
    print(f"Rijen {s['totaal']}, gevonden {s['gevonden']}, niet gevonden {s['niet_gevonden']}, "
          f"ingevuld {s['ingevuld']}, gaten {s['gaten']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
