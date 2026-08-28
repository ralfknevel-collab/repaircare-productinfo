"""
Zet het Product Data Sheet (Excel) om naar artikeldata.json: gestructureerde
productdata per artikelcode, met componenten (A/B) apart én opgeteld.

Gebruik:
    python3 ingest_artikeldata.py

Gebruikt geen API. Opnieuw draaien overschrijft artikeldata.json.
"""

from __future__ import annotations

import json
import re
import sys
import warnings
from datetime import date
from pathlib import Path

import openpyxl

LEEG = {"", "--", "-", "n.v.t.", "nvt", "inapplicable", "none"}

_GETAL = re.compile(r"-?\d+(?:[.,]\d+)?")
_PREFIX = re.compile(r"^([AB])\s*:\s*(.*)$", re.S)
_BLOK = re.compile(r"^(\d+(?:[.,]\d+)?)\s*[xX×]\s*(\d+(?:[.,]\d+)?)\s*[xX×]\s*(\d+(?:[.,]\d+)?)$")
_ROND = re.compile(r"Ø\s*:?\s*(\d+(?:[.,]\d+)?)\s*H\s*:?\s*(\d+(?:[.,]\d+)?)")


def normaliseer_kop(tekst) -> str:
    """Vouw alle whitespace samen tot één spatie en trim."""
    return " ".join(str(tekst).split()) if tekst is not None else ""


def alleen_cijfers(tekst) -> str:
    return re.sub(r"\D", "", str(tekst)) if tekst is not None else ""


def split_prefix(tekst: str) -> tuple[str | None, str]:
    """Haal een componentprefix 'A:'/'B:' van de tekst af."""
    schoon = normaliseer_kop(tekst)
    m = _PREFIX.match(schoon)
    if not m:
        return None, schoon
    return m.group(1), normaliseer_kop(m.group(2))


def parse_getal(tekst) -> float | None:
    """Eerste getal in de tekst als float; komma als decimaalteken toegestaan."""
    if tekst is None:
        return None
    if isinstance(tekst, (int, float)):
        return float(tekst)
    _, rest = split_prefix(str(tekst))
    if rest.lower() in LEEG:
        return None
    m = _GETAL.search(rest)
    if not m:
        return None
    return float(m.group(0).replace(",", "."))


def parse_maat(tekst) -> dict | None:
    """'262x290x242' -> blok; 'Ø: 48 H: 184' -> rond (l = b = diameter)."""
    if tekst is None:
        return None
    _, schoon = split_prefix(str(tekst))
    if schoon.lower() in LEEG:
        return None
    m = _BLOK.match(schoon)
    if m:
        l, b, h = (float(x.replace(",", ".")) for x in m.groups())
        return {"vorm": "blok", "l": l, "b": b, "h": h}
    m = _ROND.search(schoon)
    if m:
        d, h = (float(x.replace(",", ".")) for x in m.groups())
        return {"vorm": "rond", "diameter": d, "hoogte": h, "l": d, "b": d, "h": h}
    return None


def combineer_maat(maten: list[dict]) -> dict | None:
    """Maat van een verkoopeenheid uit componentmaten: naast elkaar gezet."""
    if not maten:
        return None
    if len(maten) == 1:
        return dict(maten[0])
    delen = " + ".join(f"{m['l']:g}" for m in maten)
    return {
        "vorm": "samengesteld",
        "l": sum(m["l"] for m in maten),
        "b": max(m["b"] for m in maten),
        "h": max(m["h"] for m in maten),
        "regel": f"componenten naast elkaar: L = {delen}, B en H = grootste component",
    }


BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "Product Data Sheet december 2024.xlsx"
UITVOER = BASE_DIR / "artikeldata.json"
KOPREGEL = 2  # 0-based: rij 3 in Excel

# Genormaliseerde kop in het sheet -> interne sleutel.
KOLOMMEN = {
    "Artikelcode": "artikelcode",
    "Omschrijving": "omschrijving",
    "Language version": "taalversie",
    "EAN-code": "ean_1",
    "Status": "status",
    "Productgroup": "productgroep",
    "Minimale Verkoop- hoeveel-heid": "min_verkoophoeveelheid",
    "Assembled item": "samengesteld",
    "Components": "component",
    "Contentes": "inhoud",
    "Dimensions per piece (mm) (LxBxH)": "maat_stuk",
    "Afmetingen collo (mm) (LxBxH)": "maat_collo",
    "Afmetingen omdoos (cm) (LxBXH)": "maat_omdoos",
    "UFI-code": "ufi",
    "VOC-content": "voc",
    "VOC-category": "voc_categorie",
    "Klasse": "klasse",
    "UN-code": "un_code",
    "Verpakkings-categorie": "verpakkingsgroep",
    "Transport-categorie ADR": "transportcategorie",
    "Dangerous for the environment": "milieugevaarlijk",
    "Flashpoint": "vlampunt",
    "Transport-naam ADR": "adr_naam",
    "GN-code": "gn_code",
    "Tarief invoer-rechten": "invoerrechten",
    "Netto gewicht per stuk (gr)": "netto_g",
    "Bruto gewicht per stuk (gr)": "bruto_g",
}
VERPLICHT = ["Artikelcode", "Omschrijving", "EAN-code", "Components",
             "Dimensions per piece (mm) (LxBxH)", "GN-code",
             "Netto gewicht per stuk (gr)", "Bruto gewicht per stuk (gr)"]

# Velden die per component (A/B) kunnen voorkomen, herkenbaar aan een 'A:'/'B:'-prefix
# of doordat ze op een componentrij staan.
COMPONENTVELDEN = {"inhoud", "maat_stuk", "ufi", "voc", "voc_categorie", "klasse", "un_code",
                   "verpakkingsgroep", "transportcategorie", "milieugevaarlijk", "vlampunt",
                   "adr_naam", "netto_g", "bruto_g"}
GETALVELDEN = {"netto_g", "bruto_g", "min_verkoophoeveelheid", "invoerrechten"}
MAATVELDEN = {"maat_stuk": "maat_mm", "maat_collo": "collo_mm", "maat_omdoos": "omdoos_cm"}
TEKSTVELDEN = COMPONENTVELDEN - GETALVELDEN - {"maat_stuk"}


def _componentnaam(cel) -> str | None:
    """Kolom 'Components' -> componentnaam, of None als de cel leeg is (bv. '--')."""
    if cel is None:
        return None
    tekst = normaliseer_kop(cel)
    if tekst.lower() in LEEG:
        return None
    return tekst


def _component(artikel: dict, naam: str) -> dict:
    for c in artikel["componenten"]:
        if c["naam"] == naam:
            return c
    c = {"naam": naam, "ruw": {}}
    artikel["componenten"].append(c)
    return c


def _zet(doel: dict, sleutel: str, tekst: str) -> None:
    """Zet een geparste waarde op artikel- of componentniveau."""
    if sleutel in GETALVELDEN:
        g = parse_getal(tekst)
        if g is not None:
            doel[sleutel] = g
    elif sleutel in MAATVELDEN:
        m = parse_maat(tekst)
        if m is not None:
            doel[MAATVELDEN[sleutel]] = m
    elif sleutel == "gn_code":
        cijfers = alleen_cijfers(tekst)
        if cijfers:
            doel[sleutel] = cijfers
    elif sleutel in TEKSTVELDEN or sleutel in ("status", "productgroep", "taalversie",
                                               "samengesteld", "omschrijving"):
        if tekst.lower() not in LEEG:
            doel[sleutel] = tekst


def _verwerk_rij(rij, index: dict[str, int], ghs_index: dict[str, int], kop_per_index: dict[int, str],
                 artikel: dict, standaard_component: str | None, hoofdrij: bool) -> None:
    """Eén sheetrij verwerken. Alle kolommen komen in 'ruw'; bekende kolommen ook geparst.

    Componentvelden met een A:/B:-prefix gaan naar dat component; zonder prefix naar het
    component van de rij (kolom Components) of, op de hoofdrij zonder component, naar het artikel.
    Overige kolommen horen op de hoofdrij bij het artikel en op een componentrij bij het component.
    """
    sleutel_per_index = {i: s for s, i in index.items()}
    ghs_kolommen = set(ghs_index.values())
    for i, kop in kop_per_index.items():
        sleutel = sleutel_per_index.get(i)
        if sleutel in ("artikelcode", "ean_1", "component") or i in ghs_kolommen:
            continue
        ruw = rij[i] if i < len(rij) else None
        if ruw is None:
            continue
        tekst = normaliseer_kop(ruw)
        if not tekst:
            continue
        if sleutel in COMPONENTVELDEN:
            prefix, rest = split_prefix(tekst)
            naam = prefix or standaard_component
            doel = _component(artikel, naam) if naam else artikel
            doel["ruw"][kop] = tekst
            _zet(doel, sleutel, rest)
        else:
            doel = artikel if hoofdrij or not standaard_component else _component(artikel, standaard_component)
            doel["ruw"][kop] = tekst
            if sleutel:
                _zet(artikel, sleutel, tekst)
    # GHS-markeringen ('x') horen bij het component van de rij, anders bij het artikel.
    ghs = [code for code, i in ghs_index.items()
           if i < len(rij) and rij[i] is not None and normaliseer_kop(rij[i])]
    if ghs:
        doel = _component(artikel, standaard_component) if standaard_component else artikel
        doel["ghs"] = ghs


def _rond_af(artikel: dict) -> None:
    """Artikelniveau afleiden uit componenten: gewichten optellen, maten combineren."""
    comps = artikel["componenten"]
    for veld, regel in (("netto_g", "netto_regel"), ("bruto_g", "bruto_regel")):
        if veld not in artikel and comps and all(veld in c for c in comps):
            artikel[veld] = sum(c[veld] for c in comps)
            artikel[regel] = "som van " + " + ".join(f"{c['naam']} {c[veld]:g} g" for c in comps)
    if "maat_mm" not in artikel:
        m = combineer_maat([c["maat_mm"] for c in comps if "maat_mm" in c])
        if m:
            artikel["maat_mm"] = m
    if "ghs" not in artikel and comps:
        gezien: list[str] = []
        for c in comps:
            for code in c.get("ghs", []):
                if code not in gezien:
                    gezien.append(code)
        if gezien:
            artikel["ghs"] = gezien


def lees_artikelen(ws) -> tuple[dict[str, dict], list[str]]:
    rijen = list(ws.iter_rows(values_only=True))
    koppen = [normaliseer_kop(k) if k is not None else None for k in rijen[KOPREGEL]]
    ontbreekt = [k for k in VERPLICHT if k not in koppen]
    if ontbreekt:
        raise ValueError(f"Verwachte kolommen ontbreken in het sheet: {', '.join(ontbreekt)}")
    index = {KOLOMMEN[k]: i for i, k in enumerate(koppen) if k in KOLOMMEN}
    ghs_index = {k: i for i, k in enumerate(koppen) if k and k.startswith("GHS")}
    kop_per_index = {i: k for i, k in enumerate(koppen) if k}
    ean_i = index["ean_1"]
    comp_i = index["component"]
    ruwe_kolommen = [k for k in koppen if k]

    artikelen: dict[str, dict] = {}
    huidig: dict | None = None
    for rij in rijen[KOPREGEL + 1:]:
        if not any(c is not None for c in rij):
            continue
        code = rij[index["artikelcode"]]
        if code is not None:
            code_str = normaliseer_kop(code)
            if isinstance(code, float) and code.is_integer():
                code_str = str(int(code))
            huidig = {"artikelcode": code_str, "omschrijving": "", "componenten": [], "ruw": {}}
            artikelen[code_str] = huidig
            ean = alleen_cijfers(rij[ean_i]) + alleen_cijfers(rij[ean_i + 1] if ean_i + 1 < len(rij) else None)
            if ean:
                huidig["ean"] = ean
            comp = _componentnaam(rij[comp_i])
            _verwerk_rij(rij, index, ghs_index, kop_per_index, huidig, comp, hoofdrij=True)
        elif huidig is not None:
            comp = _componentnaam(rij[comp_i])
            _verwerk_rij(rij, index, ghs_index, kop_per_index, huidig, comp, hoofdrij=False)
    for artikel in artikelen.values():
        _rond_af(artikel)
    return artikelen, ruwe_kolommen


def bouw_artikeldata(pad: Path) -> dict:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(pad, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    artikelen, ruwe_kolommen = lees_artikelen(ws)
    return {
        "bron": pad.name,
        "gemaakt_op": date.today().isoformat(),
        "ruwe_kolommen": ruwe_kolommen,
        "artikelen": artikelen,
    }


def main() -> int:
    if not EXCEL_FILE.exists():
        print(f"Excel niet gevonden: {EXCEL_FILE.name}")
        return 1
    data = bouw_artikeldata(EXCEL_FILE)
    UITVOER.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"{len(data['artikelen'])} artikelen geschreven naar {UITVOER.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
