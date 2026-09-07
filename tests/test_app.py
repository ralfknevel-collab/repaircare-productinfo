"""Controleer de dealerinterface zonder externe API-aanroepen."""

import io
from pathlib import Path
from types import SimpleNamespace

import pytest
import openpyxl
from streamlit.testing.v1 import AppTest

import app


def _start_app():
    import app

    app.main()


@pytest.fixture
def dealer_app(monkeypatch, tmp_path):
    monkeypatch.setattr(app.dealer_profielen, "PROFIEL_MAP", tmp_path / "profielen")
    monkeypatch.setattr(app, "get_secret", lambda naam: None)
    monkeypatch.setattr(app, "KENNISBANK_FILE", tmp_path / "kennisbank.json", raising=False)
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: SimpleNamespace())

    def geen_api(*args, **kwargs):
        raise AssertionError("Deze interfacecontrole mag geen API-client starten.")

    monkeypatch.setattr(app.anthropic, "Anthropic", geen_api)
    monkeypatch.setattr(app.anthropic, "AsyncAnthropic", geen_api)
    return AppTest.from_function(_start_app)


@pytest.fixture
def downloads(monkeypatch):
    uit = []
    origineel = app.st.download_button

    def onthoud_download(label, data, **kwargs):
        uit.append(data)
        return origineel(label, data=data, **kwargs)

    monkeypatch.setattr(app.st, "download_button", onthoud_download)
    return uit


@pytest.fixture
def doosgewicht_app(dealer_app, monkeypatch, artikeldata_dict):
    """Eén echt herkenbaar doosgewicht met dezelfde bronstructuur als het PDS."""
    from artikeldata import Artikeldata

    artikeldata_dict["artikelen"]["2023205"] = {
        "artikelcode": "2023205", "omschrijving": "DRY FLEX 1 2-in-1 (150 ml)",
        "ean": "8714748004917", "gn_code": "32141010", "netto_g": 172, "min_verkoophoeveelheid": 10,
        "ruw": {
            "Netto gewicht per doos (kg)": "A: 1,11",
            "Bruto gewicht per doos (kg)": "2.87",
            "Inhoud (om)doos": "1 x 10",
        },
        "componenten": [
            {"naam": "A", "netto_g": 111, "ruw": {"Netto gewicht per stuk (gr)": "A: 111"}},
            {"naam": "B", "netto_g": 61, "ruw": {
                "Netto gewicht per stuk (gr)": "B: 61", "Netto gewicht per doos (kg)": "B: 0,61",
            }},
        ],
    }
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vorlage Lieferanten"
    ws.append([
        "Lieferanten \nArtikelnummer", "VPE \nAnzahl", "VPE \nEinheit",
        "VPE \nBrutto-Gewicht", "VPE \nBrutto-Gewicht\nEinheit",
        "VPE \nNetto-Gewicht", "VPE \nNetto-Gewicht\nEinheit", "GN-code",
    ])
    ws.append([2023205, 10, "Karton", 2.87, "Kg", None, "Kg", None])
    inhoud = io.BytesIO()
    wb.save(inhoud)
    bestand = SimpleNamespace(name="doosgewichten.xlsx", getvalue=inhoud.getvalue)
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    return dealer_app


def test_doosgewicht_wordt_direct_ingevuld_met_componenten_en_kg(doosgewicht_app, downloads):
    at = doosgewicht_app.run()
    assert not at.exception and downloads
    wb = openpyxl.load_workbook(io.BytesIO(downloads[-1]))
    assert wb.active["F2"].value == 1.72
    assert wb.active["D2"].value == 2.87
    assert wb.active["G2"].value == "Kg"
    assert not any(veld.label == "Gewichten" for veld in at.selectbox)
    assert not any(knop.label == "Invullen" for knop in at.button)
    controle = " ".join(str(c.value) for rij in wb["Controle"] for c in rij if c.value is not None)
    assert "Product Data Sheet" in controle
    assert "1,11" in controle or "1.11" in controle
    assert "0,61" in controle or "0.61" in controle


def test_eerder_gemaakt_doosresultaat_wordt_opnieuw_verwerkt(doosgewicht_app, downloads):
    import json

    at = doosgewicht_app.run()
    assert not at.exception
    staat = at.session_state["dealer"]
    mapping = staat["mapping"].naar_dict()
    staat["uitvoersleutel"] = json.dumps(
        [deel for deel in json.loads(staat["uitvoersleutel"]) if deel != "doosgewichten_v1"],
        ensure_ascii=False, sort_keys=True,
    )
    oud = openpyxl.load_workbook(io.BytesIO(downloads[-1]))
    oud.active["F2"] = None
    buffer = io.BytesIO()
    oud.save(buffer)
    staat["uit"] = buffer.getvalue()
    at.run()
    assert not at.exception
    assert at.session_state["dealer"]["mapping"].naar_dict() == mapping
    assert openpyxl.load_workbook(io.BytesIO(downloads[-1])).active["F2"].value == 1.72


@pytest.mark.parametrize("handmatig_overslaan", [False, True])
def test_oude_dooskoppeling_vernieuwt_zonder_handmatige_keuzes_te_verliezen(
    doosgewicht_app, monkeypatch, downloads, handmatig_overslaan,
):
    import json

    oorspronkelijke_mapping = app.bepaal_mapping

    def oude_mapping(*args, **kwargs):
        mapping = oorspronkelijke_mapping(*args, **kwargs)
        for kolom in mapping.kolommen:
            if kolom.doelveld in {"collo_netto_gewicht", "collo_bruto_gewicht"}:
                kolom.doelveld, kolom.eenheid = "geen", None
        return mapping

    monkeypatch.setattr(app, "bepaal_mapping", oude_mapping)
    at = doosgewicht_app.run()
    assert not at.exception
    assert openpyxl.load_workbook(io.BytesIO(downloads[-1])).active["F2"].value is None
    # Simuleer bewerkingen die al in de open dealerinterface aanwezig zijn.
    rijen = {"7": {"Doelveld": "Niet invullen  [geen]"}}
    if handmatig_overslaan:
        rijen["5"] = {"Doelveld": "Niet invullen  [geen]"}
    wijzigingen = {"edited_rows": rijen, "added_rows": [], "deleted_rows": []}

    def herhaal_met_bewerkingen():
        # AppTest ondersteunt data_editor nog niet als widget; de browser stuurt deze toestand wel mee.
        widgets = at._tree.get_widget_states()
        widgets.widgets.add(id=at.dataframe[0].proto.id, string_value=json.dumps(wijzigingen))
        at._run(widgets)

    herhaal_met_bewerkingen()
    assert not at.exception
    assert openpyxl.load_workbook(io.BytesIO(downloads[-1])).active["H2"].value is None
    at.session_state["dealer"].pop("mapping_versie", None)
    monkeypatch.setattr(app, "bepaal_mapping", oorspronkelijke_mapping)
    herhaal_met_bewerkingen()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["F2"].value == (None if handmatig_overslaan else 1.72)
    assert ws["H2"].value is None
    eerste_resultaat = downloads[-1]
    herhaal_met_bewerkingen()
    assert not at.exception and downloads[-1] == eerste_resultaat


def test_upload_vult_direct_in_zonder_ai_of_invulknop(dealer_app, monkeypatch, artikeldata_dict, downloads):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app, "get_secret", lambda naam: "test" if naam == "ANTHROPIC_API_KEY" else None)
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: (
        b"HerstellerArtNr;GN-code;Nettogewicht (g)\n2010005;;\n2511105;behouden;\n2040005;;\n"
    ))
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception
    assert downloads
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["B2"].value == "32141010"
    assert ws["C2"].value == 318
    assert ws["B3"].value == "behouden"
    assert ws["C3"].value == 452
    assert ws["B4"].value is None and ws["C4"].value is None
    assert not any(knop.label == "Invullen" for knop in at.button)
    geavanceerd = next(v for v in at.expander if v.label == "Geavanceerd")
    assert not geavanceerd.proto.expanded
    assert any("2040005" in melding.value for melding in at.warning)


def test_onbekend_artikel_geeft_direct_eerlijk_downloadresultaat(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="onbekend.csv", getvalue=lambda: b"HerstellerArtNr;GN-code;Omschrijving\n2040005;;\n")
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception and not at.error
    assert len(downloads) == 1
    ws = openpyxl.load_workbook(io.BytesIO(downloads[0])).active
    assert ws["B2"].value is None and ws["C2"].value is None
    assert not at.success
    assert any("2040005" in melding.value for melding in at.warning)


def test_eenheidskeuze_vult_werkelijk_gewicht_en_maten_in(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="maten.csv", getvalue=lambda: (
        "HerstellerArtNr;GN-code;Nettogewicht;Länge;Breite;Höhe\n2511105;;;;;\n4513032;;;;;\n2010005;;;;;\n"
    ).encode())
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception and downloads
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["B2"].value == "32141010"
    assert ws["C2"].value is None and ws["D2"].value is None
    assert any(veld.label == "Maten" for veld in at.selectbox)
    next(veld for veld in at.selectbox if veld.label == "Maten").select("mm").run()
    next(veld for veld in at.selectbox if veld.label == "Gewichten").select("g").run()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert [ws.cell(2, kol).value for kol in range(3, 7)] == [452, 49, 49, 230]
    assert [ws.cell(3, kol).value for kol in range(3, 7)] == [120, 25, 50, 222]
    assert ws["C4"].value == 318
    assert [ws.cell(4, kol).value for kol in range(4, 7)] == [
        "A: 48 / B: 41", "A: 48 / B: 41", "A: 184 / B: 145",
    ]
    # Een andere eenheid verandert de getallen, niet de bron of artikelkoppeling.
    next(veld for veld in at.selectbox if veld.label == "Maten").select("cm").run()
    next(veld for veld in at.selectbox if veld.label == "Gewichten").select("kg").run()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert [ws.cell(2, kol).value for kol in range(3, 7)] == [0.452, 4.9, 4.9, 23]
    assert [ws.cell(4, kol).value for kol in range(3, 7)] == [
        0.318, "A: 4.8 / B: 4.1", "A: 4.8 / B: 4.1", "A: 18.4 / B: 14.5",
    ]


def test_onthouden_eenheden_vullen_volgend_bestand_zonder_extra_keuze(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="eerste.csv", getvalue=lambda: (
        "HerstellerArtNr;Primärlieferant;Nettogewicht;Länge\n2511105;973184;;\n"
    ).encode())
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception
    next(v for v in at.selectbox if v.label == "Maten").select("mm").run()
    next(v for v in at.selectbox if v.label == "Gewichten").select("g").run()
    next(k for k in at.button if k.label == "Onthouden voor dit dealerformaat").click().run()
    assert not at.exception
    bestand.name = "volgende levering.csv"
    bestand.getvalue = lambda: (
        "HerstellerArtNr;Primärlieferant;Nettogewicht;Länge\n4513032;973184;;\n"
    ).encode()
    nieuw = AppTest.from_function(_start_app).run()
    assert not nieuw.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["C2"].value == 120 and ws["D2"].value == 25
    assert not next(v for v in nieuw.expander if v.label == "Eenheden wijzigen").proto.expanded
    assert any("Onthouden voor dit dealerformaat" in t.value for t in nieuw.caption)
    bestand.getvalue = lambda: (
        "HerstellerArtNr;Primärlieferant;Nettogewicht;Länge\n4513032;999999;;\n"
    ).encode()
    ander = AppTest.from_function(_start_app).run()
    assert not ander.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["C2"].value is None and ws["D2"].value is None


def test_oude_download_wordt_vernieuwd_voor_componentmaten(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    import json
    from unittest.mock import Mock

    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="sets.csv", getvalue=lambda: (
        "HerstellerArtNr;Länge (mm)\n2010005;\n"
    ).encode())
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception
    staat = at.session_state["dealer"]
    mapping = staat["mapping"].naar_dict()
    # Een open sessie kan nog een download uit de vorige invulwerkwijze bevatten.
    staat["uitvoersleutel"] = json.dumps(
        [deel for deel in json.loads(staat["uitvoersleutel"]) if deel != "componentmaten_v1"],
        ensure_ascii=False, sort_keys=True,
    )
    oude_download = openpyxl.load_workbook(io.BytesIO(downloads[-1]))
    oude_download.active["B2"] = None
    buffer = io.BytesIO()
    oude_download.save(buffer)
    staat["uit"] = buffer.getvalue()
    verwerking = Mock(wraps=app.verwerk)
    monkeypatch.setattr(app, "verwerk", verwerking)
    at.run()
    assert not at.exception
    assert verwerking.call_count == 1
    assert at.session_state["dealer"]["mapping"].naar_dict() == mapping
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["B2"].value == "A: 48 / B: 41"


def test_onleesbaar_profiel_toont_keuze_en_verwerkt_overige_gegevens(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    inhoud = "HerstellerArtNr;GN-code;Länge\n2511105;;\n".encode()
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: inhoud)
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    ws = app.laad_werkboek(inhoud, bestand.name).active
    profiel_id = app.dealer_profielen.profielsleutel(ws, 0, bestand.name)
    # Een map op de plek van het JSON-bestand veroorzaakt een echte leesfout.
    (app.dealer_profielen.PROFIEL_MAP / f"{profiel_id}.json").mkdir(parents=True)
    at = dealer_app.run()
    assert not at.exception and downloads
    assert any("niet worden gelezen" in w.value for w in at.warning)
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["B2"].value == "32141010" and ws["C2"].value is None
    assert next(v for v in at.selectbox if v.label == "Maten").value is None


def test_mislukte_opslag_verliest_actuele_keuze_niet(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: "HerstellerArtNr;Länge\n2511105;\n".encode())
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    next(v for v in at.selectbox if v.label == "Maten").select("mm").run()

    def opslaan_geweigerd(*args):
        raise PermissionError("Geen schrijfrechten")

    monkeypatch.setattr(app.dealer_profielen.os, "replace", opslaan_geweigerd)
    next(k for k in at.button if k.label == "Onthouden voor dit dealerformaat").click().run()
    assert not at.exception
    assert any("kon niet worden opgeslagen" in w.value for w in at.warning)
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["B2"].value == 49
    nieuw = AppTest.from_function(_start_app).run()
    assert not nieuw.exception
    assert next(v for v in nieuw.selectbox if v.label == "Maten").value is None


def test_tegenstrijdige_artikelcodes_worden_niet_als_ontbrekend_voorgesteld(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="dubbel.csv", getvalue=lambda: (
        b"HerstellerArtNr;HerstellerArtNr;GN-code\n2010005;2511105;\n"
    ))
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception and downloads
    assert not any("Niet gevonden in de productbron" in melding.value for melding in at.warning)
    assert any("tegenstrijdige" in melding.value for melding in at.warning)
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["C2"].value is None


def test_tabblad_en_bestandswisseling_verversen_download(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    wb = openpyxl.Workbook()
    wb.active.title = "Eerste"
    for ws, code in [(wb.active, "2010005"), (wb.create_sheet("Tweede"), "2511105")]:
        ws.append(["HerstellerArtNr", "Nettogewicht (g)"])
        ws.append([code, None])
    inhoud = io.BytesIO()
    wb.save(inhoud)
    bestand = SimpleNamespace(name="dealer.xlsx", getvalue=inhoud.getvalue)
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception
    eerste = openpyxl.load_workbook(io.BytesIO(downloads[-1]))
    assert eerste["Eerste"]["B2"].value == 318
    assert eerste["Tweede"]["B2"].value is None
    next(veld for veld in at.selectbox if veld.label == "Tabblad").select("Tweede").run()
    assert not at.exception
    tweede = openpyxl.load_workbook(io.BytesIO(downloads[-1]))
    assert tweede["Eerste"]["B2"].value is None
    assert tweede["Tweede"]["B2"].value == 452
    bestand.name = "nieuw.csv"
    bestand.getvalue = lambda: b"HerstellerArtNr;GN-code\n4513032;\n"
    at.run()
    assert not at.exception
    nieuw = openpyxl.load_workbook(io.BytesIO(downloads[-1]))
    assert nieuw.active["B2"].value == "82055910"
    assert "Tweede" not in nieuw.sheetnames


def test_ongeldige_eenheid_verwijdert_oude_download(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    import json
    from artikeldata import Artikeldata

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: b"HerstellerArtNr;Nettogewicht (g)\n2010005;\n")
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()
    assert not at.exception and len(downloads) == 1
    bewerkingen = {"edited_rows": {"1": {"Eenheid": "cm"}}, "added_rows": [], "deleted_rows": []}
    staat = at._tree.get_widget_states()
    staat.widgets.add(id=at.dataframe[0].proto.id, string_value=json.dumps(bewerkingen))
    at._run(staat)
    assert not at.exception
    assert len(downloads) == 1
    assert at.session_state["dealer"]["uit"] is None
    assert at.warning and not at.get("download_button")


def test_bijgewerkte_productbron_en_vaste_waarden_verversen_download(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    from artikeldata import Artikeldata
    from mapping import KolomMapping, Mapping

    vaste = {"leverancier_naam": {"label": "Naam leverancier", "standaard": "Repair Care"}}
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict, vaste))
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: b"HerstellerArtNr;GN-code;Leverancier\n2010005;;\n")
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    mapping = Mapping(0, [
        KolomMapping("HerstellerArtNr", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping("GN-code", "gn_code", None, "hoog", ""),
        KolomMapping("Leverancier", "vast:leverancier_naam", None, "hoog", ""),
    ], data_start_index=1)
    monkeypatch.setattr(app, "bepaal_mapping", lambda *args: mapping)
    at = dealer_app.run()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["B2"].value == "32141010" and ws["C2"].value == "Repair Care"
    artikeldata_dict["artikelen"]["2010005"]["gn_code"] = "99999999"
    at.run()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["B2"].value == "99999999"
    vaste["leverancier_naam"]["standaard"] = "Repair Care International"
    at.run()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(downloads[-1])).active
    assert ws["C2"].value == "Repair Care International"


@pytest.mark.parametrize("voorsteltype", ["leeg", "zonder_sleutel", "onzekere_sleutel", "ongeldige_eenheid"])
def test_onbruikbaar_ai_voorstel_behoudt_bestaand_resultaat(
    dealer_app, monkeypatch, artikeldata_dict, downloads, voorsteltype,
):
    from artikeldata import Artikeldata
    from mapping import KolomMapping, Mapping

    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    monkeypatch.setattr(app, "get_secret", lambda naam: "test" if naam == "ANTHROPIC_API_KEY" else None)
    monkeypatch.setattr(app.anthropic, "AsyncAnthropic", lambda **kwargs: SimpleNamespace())
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: b"HerstellerArtNr;GN-code\n2010005;\n")
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    oorspronkelijke_mapping = app.bepaal_mapping

    def voorstel(client, *args, **kwargs):
        if client is None:
            return oorspronkelijke_mapping(client, *args, **kwargs)
        kolommen = []
        if voorsteltype != "leeg":
            kolommen.append(KolomMapping("GN-code", "gn_code", None, "hoog", ""))
        if voorsteltype == "onzekere_sleutel":
            kolommen.append(KolomMapping("HerstellerArtNr", "sleutel_artikelcode", None, "laag", ""))
        if voorsteltype == "ongeldige_eenheid":
            kolommen = [
                KolomMapping("HerstellerArtNr", "sleutel_artikelcode", None, "hoog", ""),
                KolomMapping("GN-code", "netto_gewicht", "cm", "hoog", ""),
            ]
        return Mapping(0, kolommen, data_start_index=1)

    monkeypatch.setattr(app, "bepaal_mapping", voorstel)
    at = dealer_app.run()
    assert not at.exception
    eerste = downloads[-1]
    next(knop for knop in at.button if knop.label == "AI-hulp bij kolommen").click().run()
    assert not at.exception
    assert at.warning and not at.error
    assert at.session_state["dealer"]["uit"] == eerste
    assert downloads[-1] == eerste
    assert at.session_state["dealer"]["versie"] == 0


def test_opent_dealer_zonder_kennisbank_of_api_key(dealer_app):
    at = dealer_app.run()

    assert not at.exception
    assert not at.error
    assert len(at.get("file_uploader")) == 1
    assert any("Dealerbestanden invullen" in tekst.value for tekst in at.markdown)
    assert len(at.chat_input) == 0
    assert not any("chat" in knop.label.lower() for knop in at.button)
    assert "messages" not in at.session_state


def test_wachtwoord_beschermt_dealerinterface(dealer_app, monkeypatch):
    monkeypatch.setattr(app, "get_secret", lambda naam: "testwachtwoord" if naam == "APP_PASSWORD" else None)
    at = dealer_app.run()

    assert not at.exception
    assert len(at.get("file_uploader")) == 0
    at.text_input[0].set_value("verkeerd")
    at.button[0].click().run()
    assert at.error[0].value == "Onjuist wachtwoord."
    assert len(at.get("file_uploader")) == 0

    at.text_input[0].set_value("testwachtwoord")
    at.button[0].click().run()
    assert not at.exception
    assert len(at.get("file_uploader")) == 1


def test_ontbrekende_artikeldata_blijft_zichtbaar(dealer_app, monkeypatch):
    def ontbreekt():
        raise FileNotFoundError("artikeldata.json")

    monkeypatch.setattr(app.Artikeldata, "laad", ontbreekt)
    at = dealer_app.run()

    assert not at.exception
    assert len(at.get("file_uploader")) == 0
    assert any("artikeldata.json ontbreekt" in melding.value for melding in at.error)


def test_ai_timeout_toont_handmatige_tabel_zonder_opnieuw_te_wachten(
    dealer_app, monkeypatch, artikeldata_dict, downloads,
):
    import asyncio

    import dealer_invuller
    from artikeldata import Artikeldata

    monkeypatch.setattr(app, "get_secret", lambda naam: "test" if naam == "ANTHROPIC_API_KEY" else None)
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: b"Artikelcode;Omschrijving;GN-code\n2010005;;\n")
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    monkeypatch.setattr(app.anthropic, "AsyncAnthropic", lambda **kwargs: SimpleNamespace())
    monkeypatch.setattr(app.anthropic, "Anthropic", lambda **kwargs: SimpleNamespace())
    aanroepen = []

    def trage_herkenning(*args, **kwargs):
        aanroepen.append(1)
        raise asyncio.TimeoutError("De automatische herkenning duurde te lang.")

    monkeypatch.setattr(dealer_invuller, "vraag_mapping", trage_herkenning)
    at = dealer_app.run()
    assert not at.exception
    assert len(at.dataframe) == 1
    assert aanroepen == []
    eerste_download = downloads[-1]
    next(knop for knop in at.button if knop.label == "AI-hulp bij kolommen").click().run()
    assert not at.exception
    assert any("bestaande koppelingen" in melding.value for melding in at.warning)
    assert downloads[-1] == eerste_download
    assert at.session_state["dealer"]["mapping"].sleutels()[0].doelveld == "sleutel_artikelcode"
    at.selectbox[0].select("Omschrijving").run()
    assert not at.exception
    assert aanroepen == [1]


def test_zonder_api_key_kunnen_kolommen_handmatig_gekoppeld_worden(
    dealer_app, monkeypatch, seefelder_bestand: Path, artikeldata_dict,
):
    from artikeldata import Artikeldata

    artikeldata = Artikeldata(artikeldata_dict)
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: artikeldata)
    bestand = SimpleNamespace(name=seefelder_bestand.name, getvalue=seefelder_bestand.read_bytes)
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = dealer_app.run()

    assert not at.exception
    assert not at.error
    assert len(at.dataframe) == 1
    assert any("Kolommen koppelen" in tekst.value for tekst in at.markdown)
    assert "Doelveld" in at.dataframe[0].value.columns


def test_eerste_artikelrij_blijft_behouden_tot_invullen(dealer_app, monkeypatch, artikeldata_dict):
    from artikeldata import Artikeldata
    from mapping import KolomMapping, Mapping

    artikeldata = Artikeldata(artikeldata_dict)
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: artikeldata)
    wb = openpyxl.Workbook()
    ws = wb.active
    for rij in [
        ["Productkenmerken"],
        ["Artikelnummer", "Omschrijving", "GN-code"],
        ["MANUFACTURER_PID", "DESCRIPTION_SHORT", "CUSTOMS_TARIFF_NUMBER"],
        [],
        ["2010005", "DRY FIX UNI", None],
        ["2511105", "DRY SEAL MP wit 290 ml", None],
    ]:
        ws.append(rij)
    inhoud = io.BytesIO()
    wb.save(inhoud)
    bestand = SimpleNamespace(name="kenmerken.xlsx", getvalue=inhoud.getvalue)
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    mapping = Mapping(1, [
        KolomMapping("Artikelnummer", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping("Omschrijving", "geen", None, "hoog", ""),
        KolomMapping("GN-code", "gn_code", None, "hoog", ""),
    ], data_start_index=4)
    monkeypatch.setattr(app, "bepaal_mapping", lambda *args: mapping)
    aanroepen = []
    verwerk = app.verwerk

    def onthoud_verwerking(*args, **kwargs):
        aanroepen.append(args[2].data_start_index)
        return verwerk(*args, **kwargs)

    monkeypatch.setattr(app, "verwerk", onthoud_verwerking)
    at = dealer_app.run()

    assert not at.exception
    assert [veld.value for veld in at.number_input] == [2, 5]
    at.number_input[1].set_value(6).run()
    assert not at.exception
    assert not at.error
    assert aanroepen == [4, 5]
    assert "1 cellen aangevuld" in at.success[0].value


@pytest.mark.parametrize("artikelcode,doelveld", [
    ("2010005", "geen"),
    ("<b>onbekend</b>", "gn_code"),
    ("", "gn_code"),
])
def test_niets_aanvullen_levert_geen_misleidende_succesmelding(
    dealer_app, monkeypatch, artikeldata_dict, artikelcode, doelveld, downloads,
):
    from artikeldata import Artikeldata
    from mapping import KolomMapping, Mapping

    artikeldata = Artikeldata(artikeldata_dict)
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: artikeldata)
    bestand = SimpleNamespace(
        name="dealer.csv", getvalue=lambda: f"Artikelnummer;GN-code\n{artikelcode};\n".encode(),
    )
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    mapping = Mapping(0, [
        KolomMapping("Artikelnummer", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping("GN-code", doelveld, None, "hoog", ""),
    ], data_start_index=1)
    monkeypatch.setattr(app, "bepaal_mapping", lambda *args: mapping)
    at = dealer_app.run()

    assert not at.exception
    assert downloads
    assert not at.success
    assert any("0 cellen aangevuld" in melding.value for melding in at.info)
    if artikelcode and artikelcode != "2010005":
        assert any("Niet gevonden in de productbron" in melding.value for melding in at.warning)
    if artikelcode.startswith("<"):
        resultaat = next(tekst.value for tekst in at.warning if "Niet gevonden" in tekst.value)
        assert "&lt;b&gt;onbekend&lt;/b&gt;" in resultaat
        assert "<b>onbekend</b>" not in resultaat


def test_volledige_kolomtekst_en_gewijzigde_koppeling_blijven_bewaard(
    dealer_app, monkeypatch, artikeldata_dict,
):
    import json

    from artikeldata import Artikeldata
    from mapping import KolomMapping, Mapping

    lang_voorbeeld = "Product met een uitgebreide omschrijving. " * 20 + "EINDE VOORBEELD"
    lange_uitleg = "Deze dealer vraagt specifieke productgegevens. " * 20 + "<einde toelichting>"
    inhoud = f"Artikelnummer;Omschrijving;Extra\n2010005;{lang_voorbeeld};\n".encode()
    bestand = SimpleNamespace(name="lange_tekst.csv", getvalue=lambda: inhoud)
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: Artikeldata(artikeldata_dict))
    mapping = Mapping(0, [
        KolomMapping("Artikelnummer", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping("Omschrijving", "geen", None, "middel", lange_uitleg),
        KolomMapping("Extra", "gn_code", None, "hoog", ""),
    ], data_start_index=1)
    monkeypatch.setattr(app, "bepaal_mapping", lambda *args: mapping)
    ontvangen = []
    verwerk = app.verwerk

    def onthoud_verwerking(*args, **kwargs):
        ontvangen.append(args[2])
        return verwerk(*args, **kwargs)

    monkeypatch.setattr(app, "verwerk", onthoud_verwerking)
    at = dealer_app.run()
    assert not at.exception
    kolomconfig = json.loads(at.dataframe[0].proto.columns)
    keuzelabels = [optie["label"] for optie in kolomconfig["Doelveld"]["type_config"]["options"]]
    assert len(keuzelabels) == len(set(keuzelabels))
    assert "Omschrijving (productgegeven)" in keuzelabels
    assert "Omschrijving (bronbestand)" in keuzelabels
    assert "UFI-code (productgegeven)" in keuzelabels
    assert "UFI-code (bronbestand)" in keuzelabels
    assert all("[" not in label for label in keuzelabels)
    bewerkingen = {
        "edited_rows": {"2": {"Doelveld": "Nettogewicht per stuk  [netto_gewicht]", "Eenheid": "kg"}},
        "added_rows": [], "deleted_rows": [],
    }

    def herstart_met_bewerkingen():
        # AppTest heeft geen editorbediening, dus stuur hetzelfde bericht als de browser.
        staat = at._tree.get_widget_states()
        staat.widgets.add(id=at.dataframe[0].proto.id, string_value=json.dumps(bewerkingen))
        at._run(staat)

    editor_id = at.dataframe[0].proto.id
    at.selectbox[0].select("Omschrijving")
    herstart_met_bewerkingen()
    assert not at.exception
    assert at.dataframe[0].proto.id == editor_id
    tekst = "\n".join(veld.value for veld in at.markdown)
    assert lang_voorbeeld in tekst
    assert lange_uitleg.replace("<", "&lt;").replace(">", "&gt;") in tekst
    assert "<einde toelichting>" not in tekst
    at.selectbox[0].select("Extra")
    herstart_met_bewerkingen()
    assert at.dataframe[0].proto.id == editor_id
    assert any("Nettogewicht per stuk" in veld.value for veld in at.markdown)
    herstart_met_bewerkingen()
    assert not at.exception and not at.error
    assert len(ontvangen) == 2
    assert ontvangen[-1].data_start_index == 1
    assert ontvangen[-1].kolommen[0].doelveld == "sleutel_artikelcode"
    assert ontvangen[-1].kolommen[2].doelveld == "netto_gewicht"
    assert ontvangen[-1].kolommen[2].eenheid == "kg"
