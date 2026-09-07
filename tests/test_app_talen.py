"""De taalkeuze behoudt uploads, technische gegevens en handmatige instellingen."""

import io
import json
from types import SimpleNamespace

import openpyxl
import pytest
from streamlit.testing.v1 import AppTest

import app
from artikeldata import Artikeldata


@pytest.fixture
def talen_app(monkeypatch, tmp_path):
    for naam in ("APP_LANGUAGE", "RENDER", "REQUIRE_APP_PASSWORD"):
        monkeypatch.delenv(naam, raising=False)
    monkeypatch.setattr(app, "get_secret", lambda naam: None)
    monkeypatch.setattr(app.dealer_profielen, "PROFIEL_MAP", tmp_path / "profielen")
    ad = Artikeldata({"artikelen": {"2010005": {
        "artikelcode": "2010005", "omschrijving": "DRY FIX® UNI", "ean": "8714748004368",
        "netto_g": 318, "gn_code": "32141010",
    }}})
    ad.bron_meldingen = ["Artikel 9001411: geen EAN", "Bevestigd door Ralf op 2026-09-07"]
    ad.prijslijst_info = {"bron": "verkoopadviesprijzen_2026.csv", "geldig_vanaf": "2026-01-01",
                         "geldig_tot": "2026-12-31"}
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: ad)
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: (
        b"Artikelcode;Omschrijving;Nettogewicht (g);Extra\n2010005;;;\n"
    ))
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    return AppTest.from_string("import app\napp.main()")


def test_bronuitleg_en_controlelijst_staan_niet_meer_op_het_scherm(talen_app):
    at = talen_app.run()
    assert not at.exception
    zichtbaar = " ".join(v.value for groep in (at.caption, at.text, at.markdown) for v in groep)
    assert "verkoopadviesprijzen_2026.csv" not in zichtbaar
    assert "Bevestigd door Ralf" not in zichtbaar
    assert not any(e.label.startswith("Broncontrole") for e in at.expander)


def test_duitse_interface_behoudt_productgegevens(talen_app, monkeypatch):
    monkeypatch.setenv("APP_LANGUAGE", "de")
    at = talen_app.run()
    assert not at.exception and at.get("download_button")
    assert at.radio[0].value == "de"
    assert any("Händlerdateien" in m.value for m in at.markdown)
    assert any(e.label == "Erweitert" for e in at.expander)
    assert any("Zellen ergänzt" in s.value and "gefunden" in s.value for s in at.success)
    assert "herunterladen" in at.get("download_button")[0].label.lower()
    config = json.loads(at.dataframe[0].proto.columns)
    assert config["Doelveld"]["label"] == "Produktangabe"
    labels = [optie["label"] for optie in config["Doelveld"]["type_config"]["options"]]
    assert "Nicht ausfüllen" in labels
    ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
    assert ws["B2"].value == "DRY FIX® UNI" and ws["C2"].value == 318


def test_taalwissel_behoudt_handmatige_koppeling_en_download(talen_app):
    at = talen_app.run()
    assert not at.exception
    edits = {"edited_rows": {"3": {"Doelveld": "Douanetariefnummer (GN/HS-code)  [gn_code]"}},
             "added_rows": [], "deleted_rows": []}
    widgets = at._tree.get_widget_states()
    widgets.widgets.add(id=at.dataframe[0].proto.id, string_value=json.dumps(edits))
    at._run(widgets)
    assert not at.exception
    uitvoer = at.session_state["dealer"]["uit"]
    assert openpyxl.load_workbook(io.BytesIO(uitvoer)).active["D2"].value == "32141010"
    for taal in ("de", "nl"):
        at.radio[0].set_value(taal).run()
        assert not at.exception
        ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
        assert [ws.cell(2, kolom).value for kolom in (2, 3, 4)] == ["DRY FIX® UNI", 318, "32141010"]
        assert at.dataframe[0].value.iloc[3]["Doelveld"].endswith("[gn_code]")


def test_echte_upload_blijft_bewaard_bij_taalwissel(talen_app, monkeypatch):
    from streamlit.runtime.memory_uploaded_file_manager import MemoryUploadedFileManager
    from streamlit.runtime.uploaded_file_manager import UploadedFileRec

    # Gebruik de echte uploadwidget, zodat wijzigingen aan zijn identiteit zichtbaar worden.
    monkeypatch.setattr(app.st, "file_uploader", app.st._main.file_uploader)
    record = UploadedFileRec("test-upload", "dealer.csv", "text/csv",
                             b"Artikelcode;Omschrijving;Nettogewicht (g);Extra\n2010005;;;\n")
    monkeypatch.setattr(MemoryUploadedFileManager, "get_files", lambda self, session_id, file_ids: [record])
    at = talen_app.run()
    uploader_id = at.get("file_uploader")[0].proto.id

    def met_upload(widgets):
        # AppTest heeft nog geen uploaderbediening; stuur dezelfde toestand als de browser.
        staat = widgets.widgets.add(id=uploader_id)
        bestand = staat.file_uploader_state_value.uploaded_file_info.add()
        bestand.file_id, bestand.name, bestand.size = record.file_id, record.name, len(record.data)
        return widgets

    at._run(met_upload(at._tree.get_widget_states()))
    assert not at.exception and at.get("download_button")
    edits = {"edited_rows": {"3": {"Doelveld": "Douanetariefnummer (GN/HS-code)  [gn_code]"}},
             "added_rows": [], "deleted_rows": []}
    widgets = met_upload(at._tree.get_widget_states())
    widgets.widgets.add(id=at.dataframe[0].proto.id, string_value=json.dumps(edits))
    at._run(widgets)
    assert not at.exception
    uitvoer = at.session_state["dealer"]["uit"]
    assert openpyxl.load_workbook(io.BytesIO(uitvoer)).active["D2"].value == "32141010"
    for taal in ("de", "nl"):
        at.radio[0].set_value(taal)
        at._run(met_upload(at._tree.get_widget_states()))
        assert not at.exception and at.get("download_button")
        assert at.get("file_uploader")[0].proto.id == uploader_id
        ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
        assert [ws.cell(2, kolom).value for kolom in (2, 3, 4)] == ["DRY FIX® UNI", 318, "32141010"]


@pytest.mark.parametrize("omgeving", ["RENDER", "REQUIRE_APP_PASSWORD"])
def test_online_zonder_wachtwoord_geeft_geen_toegang(talen_app, monkeypatch, omgeving):
    monkeypatch.setenv(omgeving, "true")
    at = talen_app.run()
    assert not at.exception
    assert any("APP_PASSWORD" in e.value for e in at.error)
    assert not at.get("download_button") and not at.dataframe


def test_duits_inloggen_werkt_met_wachtwoord(talen_app, monkeypatch):
    monkeypatch.setenv("APP_LANGUAGE", "de")
    monkeypatch.setenv("RENDER", "true")
    monkeypatch.setattr(app, "get_secret", lambda naam: "lang-testwachtwoord" if naam == "APP_PASSWORD" else None)
    at = talen_app.run()
    assert not at.exception and not at.get("download_button")
    assert at.text_input[0].label == "Passwort"
    at.text_input[0].set_value("verkeerd")
    at.button[0].click().run()
    assert at.error[0].value == "Falsches Passwort."
    at.text_input[0].set_value("lang-testwachtwoord")
    at.button[0].click().run()
    assert not at.exception and at.get("download_button")


def test_taalwissel_behoudt_eenheidskeuze(talen_app, monkeypatch):
    bestand = SimpleNamespace(name="eenheden.csv", getvalue=lambda: (
        b"Artikelcode;Omschrijving;Nettogewicht\n2010005;;\n"
    ))
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = talen_app.run()
    assert not at.exception
    gewicht = next(s for s in at.selectbox if s.label == "Gewichten")
    gewicht.set_value("kg").run()
    assert not at.exception
    uitvoer = at.session_state["dealer"]["uit"]
    assert openpyxl.load_workbook(io.BytesIO(uitvoer)).active["C2"].value == 0.318
    at.radio[0].set_value("de").run()
    assert not at.exception
    assert next(s for s in at.selectbox if s.label == "Gewichte").value == "kg"
    assert openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active["C2"].value == 0.318
    at.radio[0].set_value("nl").run()
    assert not at.exception
    assert next(s for s in at.selectbox if s.label == "Gewichten").value == "kg"
    assert openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active["C2"].value == 0.318


def test_taalkeuze_is_per_gebruiker(talen_app):
    nl = talen_app.run()
    de = AppTest.from_string("import app\napp.main()").run()
    de.radio[0].set_value("de").run()
    nl.run()
    assert not nl.exception and not de.exception
    assert any(e.label == "Geavanceerd" for e in nl.expander)
    assert any(e.label == "Erweitert" for e in de.expander)


@pytest.mark.parametrize("wachtwoord", ["   ", 123])
def test_online_ongeldig_ingesteld_wachtwoord_blokkeert(talen_app, monkeypatch, wachtwoord):
    monkeypatch.setenv("RENDER", "true")
    monkeypatch.setattr(app, "get_secret", lambda naam: wachtwoord if naam == "APP_PASSWORD" else None)
    at = talen_app.run()
    assert not at.exception
    assert any("APP_PASSWORD" in e.value for e in at.error)
    assert not at.get("download_button") and not at.dataframe
