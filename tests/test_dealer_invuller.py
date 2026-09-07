from pathlib import Path
from types import SimpleNamespace

import csv
import io
import json
import openpyxl
import pytest

from artikeldata import Artikeldata, Waarde
from dealer_invuller import (
    CONTROLE_TAB,
    Rapport,
    bepaal_data_start,
    bepaal_mapping,
    controleer_eenheden,
    kies_tabblad,
    koppen,
    laad_werkboek,
    lees_rijen,
    maak_waarde,
    match_rijen,
    schrijf_controle,
    verwerk,
    vind_kopregel,
    vul_in,
    werkboek_naar_bytes,
)
from dealer_invuller import main as cli_main
from mapping import KolomMapping, Mapping
from tests.conftest import SEEFELDER_KOPPEN, SEEFELDER_RIJEN, maak_dealerbestand
from tests.test_mapping import _nep_client


def test_laad_xlsx(seefelder_bestand):
    wb = laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name)
    ws = kies_tabblad(wb, None)
    assert ws.title == "Sheet1"
    assert ws.cell(1, 1).value == "ArtNr"


def test_laad_csv():
    inhoud = "ArtNr;Gewicht\n2010005;\n".encode("utf-8-sig")
    wb = laad_werkboek(inhoud, "lijst.csv")
    ws = kies_tabblad(wb, None)
    assert [c.value for c in ws[1]] == ["ArtNr", "Gewicht"]
    assert ws.cell(2, 1).value == "2010005"


def test_laad_csv_cp1252():
    # Duitse ERP-exports zijn vaak Windows-1252; umlauten mogen niet verminken.
    wb = laad_werkboek("Länge;Gewicht\n1;\n".encode("cp1252"), "lijst.csv")
    ws = kies_tabblad(wb, None)
    assert [c.value for c in ws[1]] == ["Länge", "Gewicht"]


def test_laad_onbekend_formaat():
    with pytest.raises(ValueError) as e:
        laad_werkboek(b"x", "oud.xls")
    assert ".xlsx" in str(e.value)


def test_kies_tabblad_slaat_lege_over(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Leeg"
    ws2 = wb.create_sheet("Data")
    ws2.append(["ArtNr", "EAN", "Gewicht"])
    pad = tmp_path / "twee.xlsx"
    wb.save(pad)
    wb2 = laad_werkboek(pad.read_bytes(), pad.name)
    assert kies_tabblad(wb2, None).title == "Data"
    assert kies_tabblad(wb2, "Leeg").title == "Leeg"


def test_kies_tabblad_slaat_formuleblad_en_blad_zonder_kopregel_over(tmp_path):
    wb = openpyxl.Workbook()
    hulp = wb.active
    hulp.title = "Attribuutlijst"
    hulp.append(["Regel", "Kolom", "Artikelnummer", "Waarde"])
    for r in (2, 3, 4):
        hulp.append([f"=Invoerblad!A{r}", f"=Invoerblad!B{r}", f'=CONCATENATE(INDIRECT("Invoerblad!A"&A{r}))', f"=B{r}&C{r}"])
    los = wb.create_sheet("Notities")
    los.append(["alleen wat tekst"])
    los.append([1, 2, 3])
    invoer = wb.create_sheet("Invoerblad")
    invoer.append(["Unieke sleutel", None, '=IF(COUNTA(B3:B3)=COUNTA(C3:C3),"Correct","Niet Goed")'])
    invoer.append(["Artikelnummer", "GTIN", "ArtikelnummerLeverancier", "Hoogte"])
    invoer.append([575817, 8714748004382, "2022105", None])
    pad = tmp_path / "lkp.xlsx"
    wb.save(pad)
    wb2 = laad_werkboek(pad.read_bytes(), pad.name)
    assert kies_tabblad(wb2, None).title == "Invoerblad"
    assert kies_tabblad(wb2, "Attribuutlijst").title == "Attribuutlijst"


def test_kies_tabblad_valt_terug_op_eerste_blad_met_data(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Leeg"
    alleen_formules = wb.create_sheet("Formules")
    alleen_formules.append(["a", "b", "c"])
    alleen_formules.append(["=1+1", "=2+2", "=3+3"])
    alleen_formules.append(["=4+4", "=5+5", "=6+6"])
    pad = tmp_path / "f.xlsx"
    wb.save(pad)
    wb2 = laad_werkboek(pad.read_bytes(), pad.name)
    assert kies_tabblad(wb2, None).title == "Formules"


def test_lees_rijen_en_kopregel_met_voorloop(tmp_path):
    pad = maak_dealerbestand(tmp_path / "v.xlsx", ["ArtNr", "EAN", "Gewicht (kg)"], [["1", "2", None]],
                             voorloop=[["Anfrage Stammdaten"], [], ["Bitte ausfüllen", None, None]])
    ws = kies_tabblad(laad_werkboek(pad.read_bytes(), pad.name), None)
    rijen = lees_rijen(ws, 10)
    assert rijen[0][0] == "Anfrage Stammdaten"
    assert vind_kopregel(rijen) == 3


def test_vind_kopregel_geen():
    with pytest.raises(ValueError):
        vind_kopregel([[1, 2, 3], ["a", None, None]])


def test_koppen_dedup_en_leeg(tmp_path):
    pad = maak_dealerbestand(tmp_path / "k.xlsx", ["ArtNr", " Gewicht ", None, "Gewicht", "EAN"], [[1, 2, 3, 4, 5]])
    ws = kies_tabblad(laad_werkboek(pad.read_bytes(), pad.name), None)
    k = koppen(ws, 0)
    assert k == {"ArtNr": 0, "Gewicht": 1, "Kolom C": 2, "Gewicht (2)": 3, "EAN": 4}


def test_laad_csv_fallback_muteert_stdlib_niet():
    # Eén kolom: de Sniffer kan geen scheidingsteken bepalen -> fallback op ';'.
    wb = laad_werkboek("ArtNr\n2010005\n".encode("utf-8"), "een.csv")
    ws = kies_tabblad(wb, None)
    assert ws.cell(1, 1).value == "ArtNr" and ws.cell(2, 1).value == "2010005"
    assert csv.excel.delimiter == ","


SEEFELDER_MAPPING = Mapping(0, [
    KolomMapping("ArtNr", "geen", None, "hoog", "eigen nummer dealer"),
    KolomMapping("Bundesland", "vast:bundesland", None, "hoog", ""),
    KolomMapping("Ursprungsland", "vast:ursprungsland", None, "hoog", ""),
    KolomMapping("Zolltarifnummer", "gn_code", None, "hoog", ""),
    KolomMapping("Nettogewicht", "netto_gewicht", "g", "middel", ""),
    KolomMapping("Länge", "lengte", "cm", "middel", ""),
    KolomMapping("Breite", "breedte", "cm", "middel", ""),
    KolomMapping("Höhe", "hoogte", "cm", "middel", ""),
    KolomMapping("ArtBeschreibung", "geen", None, "hoog", ""),
    KolomMapping("Primärlieferant", "geen", None, "hoog", ""),
    KolomMapping("VKEinheit", "geen", None, "hoog", ""),
    KolomMapping("HerstellerArtNr", "sleutel_artikelcode", None, "hoog", ""),
    KolomMapping("EAN13", "sleutel_ean", None, "hoog", ""),
])

VASTE_TEST = {"ursprungsland": {"label": "Land", "standaard": None, "per_prefix": {"2": "NLD"}, "per_artikel": {}},
              "bundesland": {"label": "Bundesland", "standaard": None}}


@pytest.fixture
def ad(artikeldata_dict):
    return Artikeldata(artikeldata_dict, VASTE_TEST)


@pytest.mark.parametrize("w, doel, verwacht", [
    (Waarde(318.0, "g", "b"), "kg", 0.318),
    (Waarde(318.0, "g", "b"), "g", 318),
    (Waarde(184.0, "mm", "b"), "cm", 18.4),
    (Waarde(89.0, "mm", "b"), None, 89),
    (Waarde("32141010", None, "b"), None, "32141010"),
    (Waarde(0.3333333, "kg", "b"), "kg", 0.333),
])
def test_maak_waarde(w, doel, verwacht):
    assert maak_waarde(w, doel) == verwacht
    assert type(maak_waarde(w, doel)) is type(verwacht)


def test_match_rijen(seefelder_bestand, ad):
    ws = kies_tabblad(laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name), None)
    res = match_rijen(ws, SEEFELDER_MAPPING, ad)
    assert [r.rij for r in res] == [2, 3, 4, 5, 6]
    assert res[0].match.via == "artikelcode"
    assert res[3].sleutel == "0 / 8714748004955"
    assert res[3].match is None                 # Wipes zit niet in de fixture
    assert all(r.velden == [] for r in res)


def test_vul_in_seefelder(seefelder_bestand, ad):
    wb = laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name)
    ws = kies_tabblad(wb, None)
    rapport = vul_in(ws, SEEFELDER_MAPPING, ad)

    # Rij 2 = DRY FIX UNI: gn, gewicht, maten in cm, land NLD via prefix, Bundesland leeg+geel.
    assert ws["D2"].value == "32141010"
    assert ws["E2"].value == 318
    assert (ws["F2"].value, ws["G2"].value, ws["H2"].value) == (
        "A: 4.8 / B: 4.1", "A: 4.8 / B: 4.1", "A: 18.4 / B: 14.5",
    )
    assert ws["C2"].value == "NLD"
    assert ws["B2"].value is None and ws["B2"].fill.start_color.rgb.endswith("FFFF00")
    # Rij 4 = spatel: bestaande GN-code blijft staan, land leeg (prefix 4 niet geconfigureerd).
    assert ws["D4"].value == "82055910"
    assert ws["C4"].value is None
    # Rij 5 = Wipes: niet gevonden -> alle doelcellen geel, leeg.
    assert ws["D5"].value is None and ws["D5"].fill.start_color.rgb.endswith("FFFF00")
    # Rij 6 = Box: geen GN, geen maat -> geel; gewicht wel.
    assert ws["D6"].value is None and ws["E6"].value == 8710
    # 'geen'-kolommen ongemoeid.
    assert ws["I2"].value == "REPAIR CARE DRY FIX UNI"

    s = rapport.samenvatting()
    assert s["totaal"] == 5 and s["gevonden"] == 4 and s["niet_gevonden"] == 1
    assert s["via"] == {"artikelcode": 4}
    assert s["gaten_per_kolom"]["Bundesland"] == 5
    assert s["gaten_per_kolom"]["Zolltarifnummer"] == 2   # Wipes + Box
    statussen = {(v.kolom, v.status) for v in rapport.rijen[2].velden}
    assert ("Zolltarifnummer", "bestaand") in statussen


def test_vul_in_overschrijven(seefelder_bestand, ad):
    wb = laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name)
    ws = kies_tabblad(wb, None)
    vul_in(ws, SEEFELDER_MAPPING, ad, overschrijven=True)
    assert ws["D4"].value == "82055910"  # zelfde waarde uit de data, nu wél geschreven


def test_vul_in_zonder_sleutel_faalt(seefelder_bestand, ad):
    ws = kies_tabblad(laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name), None)
    m = Mapping(0, [KolomMapping("Zolltarifnummer", "gn_code", None, "hoog", "")])
    with pytest.raises(ValueError):
        vul_in(ws, m, ad)


def test_controle_tab(seefelder_bestand, ad):
    wb = laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name)
    ws = kies_tabblad(wb, None)
    rapport = vul_in(ws, SEEFELDER_MAPPING, ad)
    schrijf_controle(wb, rapport)
    schrijf_controle(wb, rapport)  # tweede keer: vervangen, niet dupliceren
    assert wb.sheetnames.count(CONTROLE_TAB) == 1
    ct = wb[CONTROLE_TAB]
    tekst = "\n".join(" ".join(str(c) for c in rij if c is not None) for rij in ct.iter_rows(values_only=True))
    assert "2010005" in tekst and "artikelcode" in tekst
    assert "222" in tekst and "96" in tekst          # rekenregel gewicht
    assert "geen totale setmaat" in tekst               # bronbetekenis van de componentmaten
    assert "componenten A en B" in tekst and "A: 4.8 / B: 4.1" in tekst
    assert "niet gevonden" in tekst.lower()
    assert "Gevonden: 4" in tekst


def test_verwerk_rondreis(seefelder_bestand, ad):
    uit, rapport = verwerk(seefelder_bestand.read_bytes(), seefelder_bestand.name, SEEFELDER_MAPPING, ad)
    wb = openpyxl.load_workbook(io.BytesIO(uit))
    assert CONTROLE_TAB in wb.sheetnames
    assert wb["Sheet1"]["E2"].value == 318
    assert rapport.samenvatting()["gevonden"] == 4


def test_verwerk_csv_met_kg_en_cm(tmp_path, ad):
    inhoud = "Item no.;Net weight (kg);Height (cm)\n2010005;;\n".encode("utf-8")
    m = Mapping(0, [KolomMapping("Item no.", "sleutel_artikelcode", None, "hoog", ""),
                    KolomMapping("Net weight (kg)", "netto_gewicht", "kg", "hoog", ""),
                    KolomMapping("Height (cm)", "hoogte", "cm", "hoog", "")])
    uit, _ = verwerk(inhoud, "lijst.csv", m, ad)
    ws = openpyxl.load_workbook(io.BytesIO(uit))["Sheet1"]
    assert ws["B2"].value == 0.318 and ws["C2"].value == "A: 18.4 / B: 14.5"


@pytest.mark.parametrize("eenheid, verwacht", [
    ("mm", "A: 48 / B: 41"),
    ("cm", "A: 4.8 / B: 4.1"),
    ("m", "A: 0.048 / B: 0.041"),
])
def test_componentmaten_worden_per_component_omgerekend(ad, eenheid, verwacht):
    waarde = ad.waarde(ad.zoek(artikelcode="2010005").artikel, "lengte")
    assert maak_waarde(waarde, eenheid) == verwacht


def test_componentmaten_maken_alleen_ingevulde_maatkolommen_breder(seefelder_bestand, ad):
    wb = laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name)
    ws = wb.active
    for kolom in ("E", "F", "G", "H", "I"):
        ws.column_dimensions[kolom].width = 6
    ws.column_dimensions["G"].width = 30
    ws["H2"] = "bestaande maat"
    stijlen = {c.coordinate: list(c._style or [0] * 9) for rij in ws for c in rij}

    vul_in(ws, SEEFELDER_MAPPING, ad, behoud_sjabloon=True)

    assert ws.column_dimensions["F"].width >= len(ws["F2"].value) + 2
    assert ws.column_dimensions["G"].width == 30
    assert ws.column_dimensions["H"].width == 6
    assert ws.column_dimensions["E"].width == 6
    assert ws.column_dimensions["I"].width == 6
    assert ws["H2"].value == "bestaande maat"
    assert {c.coordinate: list(c._style or [0] * 9) for rij in ws for c in rij} == stijlen


@pytest.mark.parametrize("maatkolom", ["B", "C"])
@pytest.mark.parametrize("breedte", [6, 40])
def test_componentmaten_behouden_gedeelde_kolominstellingen(ad, maatkolom, breedte):
    ws = openpyxl.Workbook().active
    ws.append(["Artikelcode", "Tweede", "Derde", "Vierde"])
    ws.append(["2010005", None, None, None])
    ws.column_dimensions.group("B", "D", hidden=True)
    ws.column_dimensions["B"].width = breedte
    instellingen = {naam: dict(dimensie) for naam, dimensie in ws.column_dimensions.items()}
    mapping = Mapping(0, [
        KolomMapping("Artikelcode", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping(ws[f"{maatkolom}1"].value, "lengte", "mm", "hoog", ""),
    ])

    vul_in(ws, mapping, ad, behoud_sjabloon=True)

    assert ws[f"{maatkolom}2"].value == "A: 48 / B: 41"
    assert {naam: dict(dimensie) for naam, dimensie in ws.column_dimensions.items()} == instellingen


def test_componentmaten_verkleinen_standaard_kolombreedte_niet(ad):
    ws = openpyxl.Workbook().active
    ws.append(["Artikelcode", "Lengte"])
    ws.append(["2010005", None])
    ws.sheet_format.defaultColWidth = 40
    mapping = Mapping(0, [
        KolomMapping("Artikelcode", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping("Lengte", "lengte", "mm", "hoog", ""),
    ])

    vul_in(ws, mapping, ad, behoud_sjabloon=True)

    assert ws["B2"].value == "A: 48 / B: 41"
    assert "B" not in ws.column_dimensions
    assert ws.sheet_format.defaultColWidth == 40


def test_componentmaten_rondreis_met_bron_en_bestaande_cellen(ad):
    inhoud = b"Artikelcode;Lengte;Breedte;Hoogte\n2010005;;dealermaat;=10+20\n"
    mapping = Mapping(0, [KolomMapping("Artikelcode", "sleutel_artikelcode", None, "hoog", "")] + [
        KolomMapping(kop, doel, "mm", "hoog", "Keuze gebruiker: mm.")
        for kop, doel in [("Lengte", "lengte"), ("Breedte", "breedte"), ("Hoogte", "hoogte")]
    ])
    uit, rapport = verwerk(inhoud, "sets.csv", mapping, ad, behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(uit))
    assert [wb["Sheet1"].cell(2, k).value for k in (2, 3, 4)] == [
        "A: 48 / B: 41", "dealermaat", "=10+20",
    ]
    resultaat = rapport.rijen[0].velden[0]
    assert resultaat.waarde == "A: 48 / B: 41" and resultaat.eenheid == "mm"
    assert "componenten A en B" in resultaat.bron
    assert "geen totale setmaat" in resultaat.regel and "Keuze gebruiker: mm." in resultaat.regel
    assert [v.status for v in rapport.rijen[0].velden] == ["ingevuld", "bestaand", "overgeslagen"]
    controletekst = str(list(wb[CONTROLE_TAB].values))
    assert "A: 48 / B: 41" in controletekst and "componenten A en B" in controletekst


@pytest.mark.parametrize("behoud_sjabloon", [False, True])
def test_onvolledige_componentmaat_blijft_leeg_met_duidelijke_reden(ad, behoud_sjabloon):
    artikel = ad.zoek(artikelcode="2010005").artikel
    del artikel["componenten"][1]["maat_mm"]["l"]
    wb = openpyxl.Workbook()
    wb.active.append(["Artikelcode", "Lengte", "Hoogte"])
    wb.active.append(["2010005"])
    mapping = Mapping(0, [
        KolomMapping("Artikelcode", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping("Lengte", "lengte", "mm", "hoog", ""),
        KolomMapping("Hoogte", "hoogte", "mm", "hoog", ""),
    ])
    rapport = vul_in(wb.active, mapping, ad, behoud_sjabloon=behoud_sjabloon)
    assert wb.active["B2"].value is None
    assert wb.active["C2"].value == "A: 184 / B: 145"
    leeg = rapport.rijen[0].velden[0]
    assert leeg.status == "onzeker" and "component" in leeg.regel.lower()


def test_cli_met_mapping_bestand(seefelder_bestand, tmp_path, artikeldata_dict, monkeypatch):
    pj = tmp_path / "artikeldata.json"
    pj.write_text(json.dumps(artikeldata_dict), encoding="utf-8")
    pv = tmp_path / "vaste.json"
    pv.write_text(json.dumps(VASTE_TEST), encoding="utf-8")
    import artikeldata as ad_mod
    monkeypatch.setattr(ad_mod, "ARTIKELDATA_FILE", pj)
    monkeypatch.setattr(ad_mod, "VASTE_WAARDEN_FILE", pv)

    pm = tmp_path / "mapping.json"
    pm.write_text(json.dumps(SEEFELDER_MAPPING.naar_dict()), encoding="utf-8")
    uit = tmp_path / "uit.xlsx"
    code = cli_main([str(seefelder_bestand), "--mapping", str(pm), "--uit", str(uit)])
    assert code == 0
    wb = openpyxl.load_workbook(uit)
    assert wb["Sheet1"]["E2"].value == 318
    assert CONTROLE_TAB in wb.sheetnames


def test_cli_zonder_sleutel_geeft_melding(seefelder_bestand, tmp_path, artikeldata_dict, monkeypatch, capsys):
    pj = tmp_path / "artikeldata.json"
    pj.write_text(json.dumps(artikeldata_dict), encoding="utf-8")
    import artikeldata as ad_mod
    monkeypatch.setattr(ad_mod, "ARTIKELDATA_FILE", pj)
    monkeypatch.setattr(ad_mod, "VASTE_WAARDEN_FILE", tmp_path / "geen.json")
    pm = tmp_path / "leeg.json"
    pm.write_text(json.dumps(Mapping(0, []).naar_dict()), encoding="utf-8")
    assert cli_main([str(seefelder_bestand), "--mapping", str(pm), "--uit", str(tmp_path / "u.xlsx")]) == 1
    assert "Geen sleutelkolom" in capsys.readouterr().out


def test_controleer_eenheden():
    m = Mapping(0, [
        KolomMapping("HerstellerArtNr", "sleutel_artikelcode", None, "hoog", ""),
        KolomMapping("Nettogewicht", "netto_gewicht", "cm", "middel", ""),
        KolomMapping("Länge", "lengte", "cm", "hoog", ""),
        KolomMapping("Zolltarifnummer", "gn_code", None, "hoog", ""),
    ])
    meldingen = controleer_eenheden(m)
    assert len(meldingen) == 1
    assert meldingen[0] == "Kolom 'Nettogewicht': eenheid cm past niet bij Nettogewicht per stuk (g)"

    m.kolommen[1].eenheid = "kg"
    assert controleer_eenheden(m) == []


def test_vul_in_meldt_kolommen_buiten_de_kopregel(seefelder_bestand, ad):
    ws = kies_tabblad(laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name), None)
    m = Mapping(0, SEEFELDER_MAPPING.kolommen + [
        KolomMapping("Bestaat niet", "bruto_gewicht", "g", "laag", ""),
        KolomMapping("Ook weg", "sleutel_ean", None, "laag", ""),
        KolomMapping("Genegeerd", "geen", None, "laag", ""),
    ])
    rapport = vul_in(ws, m, ad)
    assert rapport.overgeslagen_kolommen == ["Bestaat niet", "Ook weg"]


def test_schrijf_controle_meldt_overgeslagen_kolommen(seefelder_bestand, ad):
    wb = laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name)
    ws = kies_tabblad(wb, None)
    rapport = vul_in(ws, SEEFELDER_MAPPING, ad)
    rapport.overgeslagen_kolommen = ["Bestaat niet"]
    schrijf_controle(wb, rapport)
    tekst = "\n".join(" ".join(str(c) for c in rij if c is not None)
                      for rij in wb[CONTROLE_TAB].iter_rows(values_only=True))
    assert "Overgeslagen kolommen (niet in kopregel): Bestaat niet" in tekst


# --- bepaal_mapping ---------------------------------------------------------

def _antwoord_een_kolom(kolom: str) -> dict:
    return {"kopregel_index": 0, "kolommen": [
        {"kolom": kolom, "doelveld": "sleutel_artikelcode", "eenheid": "",
         "zekerheid": "hoog", "toelichting": ""}], "opmerkingen": ""}


def test_bepaal_mapping_vult_aan_en_reconcilieert_kolomnamen(seefelder_bestand, ad):
    ws = kies_tabblad(laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name), None)
    # Claude noemt de kolom met afwijkende hoofdletters en extra whitespace.
    client, aanroepen = _nep_client(_antwoord_een_kolom("  herstellerartnr "))
    m = bepaal_mapping(client, ws, ad)

    assert m.kopregel_index == 0
    assert len(aanroepen) == 1
    # Claude's kolom eerst, daarna de rest in kopregelvolgorde als 'geen'.
    assert sorted(k.kolom for k in m.kolommen) == sorted(SEEFELDER_KOPPEN)
    assert m.kolommen[0].kolom == "HerstellerArtNr" and m.kolommen[0].doelveld == "sleutel_artikelcode"
    assert all(k.doelveld == "geen" for k in m.kolommen[1:])
    assert m.opmerkingen == ""


def test_bepaal_mapping_verwijdert_onbekende_kolom(seefelder_bestand, ad):
    ws = kies_tabblad(laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name), None)
    client, _ = _nep_client(_antwoord_een_kolom("Bestaat niet"))
    m = bepaal_mapping(client, ws, ad)
    assert "Bestaat niet" not in [k.kolom for k in m.kolommen]
    assert [k.kolom for k in m.kolommen] == SEEFELDER_KOPPEN
    assert "Kolom 'Bestaat niet' uit het Claude-voorstel niet gevonden in de kopregel." in m.opmerkingen


def test_bepaal_mapping_zonder_client(seefelder_bestand, ad):
    ws = kies_tabblad(laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name), None)
    m = bepaal_mapping(None, ws, ad)
    assert [k.kolom for k in m.kolommen] == SEEFELDER_KOPPEN
    assert {k.kolom: k.doelveld for k in m.sleutels()} == {
        "HerstellerArtNr": "sleutel_artikelcode", "EAN13": "sleutel_ean",
    }
    assert "handmatig" in m.opmerkingen


def test_bepaal_mapping_bij_fout_lege_mapping(seefelder_bestand, ad, monkeypatch):
    ws = kies_tabblad(laad_werkboek(seefelder_bestand.read_bytes(), seefelder_bestand.name), None)

    def vraag(*args, **kwargs):
        raise ValueError("API stuk")

    monkeypatch.setattr("dealer_invuller.vraag_mapping", vraag)
    m = bepaal_mapping(object(), ws, ad)
    assert [k.kolom for k in m.kolommen] == SEEFELDER_KOPPEN
    assert {k.kolom for k in m.sleutels()} == {"HerstellerArtNr", "EAN13"}
    assert "API stuk" in m.opmerkingen


@pytest.mark.parametrize("fout", ["totaal", "lezen", "onderbroken"])
def test_bepaal_mapping_na_wachttijd_bewaart_sleutels_en_beginrij(ad, monkeypatch, fout):
    import asyncio
    import httpx

    fouten = {"totaal": asyncio.TimeoutError(), "lezen": httpx.ReadTimeout("Geen data"),
              "onderbroken": httpx.RemoteProtocolError("Stream afgebroken")}
    def vraag(*args, **kwargs):
        raise fouten[fout]
    monkeypatch.setattr("dealer_invuller.vraag_mapping", vraag)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Material", "Produkt", "EM", "MW"])
    ws.append(["Standard-Material", "Lief-Mat", "EAN", "Gewicht"])
    ws.append([])
    ws.append([None, None, "MATNR", "MEMD0002"])
    ws.append(["dealer-123", "2010005", "8714748004368", None])
    m = bepaal_mapping(object(), ws, ad)
    assert m.kopregel_index == 1 and m.data_start_index == 4
    assert {k.kolom for k in m.sleutels()} == {"Lief-Mat", "EAN"}
    assert not m.doelen()
    assert "handmatig" in m.opmerkingen
    if fout == "totaal":
        assert "wachttijd" in m.opmerkingen.lower()


def test_bepaal_mapping_zonder_kopregel(tmp_path, ad):
    pad = maak_dealerbestand(tmp_path / "geenkop.xlsx", [1, 2, 3], [[4, 5, 6]])
    ws = kies_tabblad(laad_werkboek(pad.read_bytes(), pad.name), None)
    m = bepaal_mapping(None, ws, ad)
    assert m.kopregel_index == 0 and m.kolommen == []
    assert "kopregel" in m.opmerkingen.lower()


@pytest.mark.parametrize("tweede_kop", ["Gewicht", "GEWICHT"])
def test_api_dubbele_koppen_worden_aan_aparte_kolommen_gekoppeld(ad, tweede_kop):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Artikelcode", "Gewicht", tweede_kop])
    ws.append(["2010005", None, None])
    client, _ = _nep_client({
        "kopregel_index": 0, "data_start_index": 1, "opmerkingen": "",
        "kolommen": [
            {"kolom": "Artikelcode", "doelveld": "sleutel_artikelcode", "eenheid": ""},
            {"kolom": "Gewicht", "doelveld": "netto_gewicht", "eenheid": "g"},
            {"kolom": tweede_kop, "doelveld": "bruto_gewicht", "eenheid": "g"},
        ],
    })
    m = bepaal_mapping(client, ws, ad)
    vul_in(ws, m, ad)
    assert [k.kolom for k in m.kolommen] == ["Artikelcode", "Gewicht", "Gewicht (2)" if tweede_kop == "Gewicht" else tweede_kop]
    assert ws["B2"].value == 318
    assert ws["C2"].value == 360


@pytest.mark.parametrize("code", ["MEMD0002", "MEMB0005", "MWMEMB0005-005"])
def test_meerlaagse_koppen_en_technische_codes(ad, code):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Material", "Produkt", "EM", "MW"])
    ws.append(["Standard-Material", "Lief-Mat", "EAN", "Gewicht"])
    ws.append([])
    ws.append([None, None, "MATNR", code])
    ws.append(["dealer-123", "2010005", "8714748004368", None])
    ws.append([None, None, None, "Toelichting voor de dealer"])
    ws.cell(30, 4).number_format = "0.0"
    assert vind_kopregel(lees_rijen(ws)) == 1
    assert bepaal_data_start(ws, 1) == 4
    m = bepaal_mapping(None, ws, ad)
    assert m.kopregel_index == 1 and m.data_start_index == 4
    assert {k.kolom for k in m.sleutels()} == {"Lief-Mat", "EAN"}
    assert [r.rij for r in match_rijen(ws, m, ad)] == [5]
    assert match_rijen(ws, m, ad)[0].match.artikel["artikelcode"] == "2010005"


def test_weigel_koppen_met_regeleinden_en_ean_punten(ad):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Lieferanten \nArtikelnummer", "Gtin\n(EAN)", "Artikelname"])
    ws.append(["2010005", "87.14748.00436.8", "DRY FIX UNI"])
    ws.append(["onbekend", "87.14748.00380.4", "DRY SEAL"])
    m = bepaal_mapping(None, ws, ad)
    assert [r.match.artikel["artikelcode"] for r in match_rijen(ws, m, ad)] == ["2010005", "2511105"]


def test_beschermde_cellen_blijven_intact_bij_overschrijven(ad):
    from openpyxl.styles import Color, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["HerstellerArtNr", "Zwart", "Formule", "Samengevoegd", "Gewicht"])
    ws.append(["2010005", None, "=10+20", None, None])
    ws["B2"].fill = PatternFill("solid", fgColor=Color(theme=1))
    ws.merge_cells("C2:D2")
    legenda = wb.create_sheet("Legende")
    legenda.append(["Bedeutung der schwarzen Zellen", "Diese Felder sollen nicht befüllt werden"])
    # Opslaan en laden geeft het werkboek het standaard Excel-kleurthema.
    wb = laad_werkboek(werkboek_naar_bytes(wb), "test.xlsx")
    m = Mapping(0, [KolomMapping("HerstellerArtNr", "sleutel_artikelcode", None, "hoog", "")] + [
        KolomMapping(k, "netto_gewicht", "g", "hoog", "")
        for k in ["Zwart", "Formule", "Samengevoegd", "Gewicht"]
    ])
    ws = wb.active
    rapport = vul_in(ws, m, ad, overschrijven=True)
    assert ws["B2"].value is None and ws["B2"].fill.fgColor.theme == 1
    assert ws["C2"].value == "=10+20" and ws["D2"].value is None
    assert ws["E2"].value == 318
    assert [v.status for v in rapport.rijen[0].velden] == ["overgeslagen"] * 3 + ["ingevuld"]


def test_start_rij_buiten_werkblad_geeft_duidelijke_fout(ad):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Artikel", "EAN", "Gewicht"])
    ws.append(["2010005"])
    m = Mapping(0, [KolomMapping("Artikel", "sleutel_artikelcode", None, "hoog", "")], data_start_index=0)
    with pytest.raises(ValueError, match="eerste artikelrij"):
        match_rijen(ws, m, ad)
