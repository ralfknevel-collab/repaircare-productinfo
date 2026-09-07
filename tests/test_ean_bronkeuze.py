"""Een bevestigde EAN-keuze geldt alleen voor het bedoelde artikel en bronpaar."""

from copy import deepcopy
from datetime import date
import io

import openpyxl
import pytest

from artikeldata import Artikeldata
from dealer_invuller import bepaal_mapping, verwerk, werkboek_naar_bytes
from mapping import KolomMapping


@pytest.fixture
def bronnen(monkeypatch):
    class Vandaag(date):
        @classmethod
        def today(cls):
            return cls(2026, 9, 7)
    monkeypatch.setattr("artikeldata.date", Vandaag)
    pds = {"artikelen": {"2023005": {
        "artikelcode": "2023005", "omschrijving": "DRY FLEX 1", "ean": "8714748002616",
        "netto_g": 340, "maat_mm": {"l": 50, "b": 50, "h": 240},
    }}}
    prijzen = {
        "bron": "verkoopadviesprijzen_2026.csv", "geldig_vanaf": "2026-01-01", "geldig_tot": "2026-12-31",
        "valuta": "EUR", "btw": "exclusief", "meldingen": [],
        "artikelen": {"2023005": {
            "omschrijving": "DRY FLEX® 1", "ean": "8714748004740", "ve_aantal": 20,
            "eenheid": "set", "adviesprijs_cent": 7458, "bronregel": 6,
        }},
    }
    return pds, prijzen


def test_bevestigde_ean_koppelt_prijslijst_en_behoudt_technische_bron(bronnen):
    pds, prijzen = bronnen
    bronkopie = deepcopy(bronnen)
    ad = Artikeldata(pds, prijslijst=prijzen)
    artikel = ad.zoek(artikelcode="2023005").artikel
    assert artikel["ean"] == "8714748004740"
    assert ad.zoek(ean="8714748004740").artikel is artikel
    assert ad.zoek(ean="8714748002616") is None
    assert not artikel.get("bron_conflicten")
    assert ad.waarde(artikel, "adviesprijs").waarde == 74.58
    assert ad.waarde(artikel, "netto_gewicht").waarde == 340
    assert ad.waarde(artikel, "lengte").waarde == 50
    assert ad.waarde(artikel, "omschrijving").waarde == "DRY FLEX® 1"
    assert ad.waarde(artikel, "netto_gewicht").bron == "Product Data Sheet"
    ean = ad.waarde(artikel, "ean")
    assert ean.waarde == "8714748004740"
    assert "verkoopadviesprijzen_2026.csv" in ean.bron and "rij 6" in ean.bron
    assert "8714748002616" in ean.regel and "bevestigd" in ean.regel.lower()
    assert any("2023005" in m and "bevestigd" in m.lower() for m in ad.bron_meldingen)
    assert bronnen == bronkopie


@pytest.mark.parametrize("wijziging", ["artikelcode", "pds_ean", "prijs_ean", "ander_artikel"])
def test_bevestiging_laat_andere_bronconflicten_niet_door(bronnen, wijziging):
    pds, prijzen = bronnen
    code = "2023005"
    if wijziging == "artikelcode":
        code = "2023999"
        pds["artikelen"][code] = pds["artikelen"].pop("2023005")
        pds["artikelen"][code]["artikelcode"] = code
        prijzen["artikelen"][code] = prijzen["artikelen"].pop("2023005")
    elif wijziging == "pds_ean":
        pds["artikelen"][code]["ean"] = "8714748004368"
    elif wijziging == "prijs_ean":
        prijzen["artikelen"][code]["ean"] = "8714748004368"
    else:
        pds["artikelen"]["2023999"] = {"artikelcode": "2023999", "ean": "8714748004740"}
    ad = Artikeldata(pds, prijslijst=prijzen)
    artikel = ad.zoek(artikelcode=code).artikel
    assert artikel.get("bron_conflicten")
    assert not ad.waarde(artikel, "adviesprijs").eenduidig
    assert artikel["ean"] == pds["artikelen"][code]["ean"]
    if wijziging == "ander_artikel":
        assert ad.zoek(ean="8714748004740").artikel["artikelcode"] == "2023999"
        assert ad.artikelen["2023999"]["bron_conflicten"]


def test_invullen_via_bevestigde_ean_en_ean_export_met_juiste_bron(bronnen):
    ad = Artikeldata(bronnen[0], prijslijst=bronnen[1])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["EAN", "Nettogewicht (g)", "Adviesprijs excl. btw (EUR)", "Omschrijving", "EAN uitvoer"])
    ws.append(["8714748004740", None, None, "Naam dealer"])
    mapping = bepaal_mapping(None, ws, ad)
    mapping.kolommen[-1] = KolomMapping("EAN uitvoer", "ean", None, "hoog", "Handmatig gekozen")
    uit, rapport = verwerk(werkboek_naar_bytes(wb), "dealer.xlsx", mapping, ad, behoud_sjabloon=True)
    resultaat = openpyxl.load_workbook(io.BytesIO(uit))
    assert [resultaat.active.cell(2, c).value for c in range(1, 6)] == [
        "8714748004740", 340, 74.58, "Naam dealer", "8714748004740",
    ]
    assert not rapport.rijen[0].toelichting
    assert rapport.samenvatting()["ingevuld"] == 3
    controle = " ".join(str(c.value) for rij in resultaat["Controle"] for c in rij if c.value is not None)
    assert "verkoopadviesprijzen_2026.csv" in controle and "8714748002616" in controle


def test_open_app_vernieuwt_geblokkeerde_download_na_bronkeuze(bronnen, monkeypatch, tmp_path):
    from types import SimpleNamespace
    from streamlit.testing.v1 import AppTest
    import app

    pds, prijzen = bronnen
    botsende_bron = deepcopy(pds)
    botsende_bron["artikelen"]["2023999"] = {"artikelcode": "2023999", "ean": "8714748004740"}
    huidig = {"data": Artikeldata(botsende_bron, prijslijst=prijzen)}
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: huidig["data"])
    monkeypatch.setattr(app, "get_secret", lambda naam: None)
    monkeypatch.setattr(app.dealer_profielen, "PROFIEL_MAP", tmp_path / "profielen")
    bestand = SimpleNamespace(name="bronkeuze.csv", getvalue=lambda: (
        b"Artikelcode;Nettogewicht (g);Adviesprijs excl. btw (EUR)\n2023005;;\n"
    ))
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)
    at = AppTest.from_string("import app\napp.main()").run()
    assert not at.exception
    assert openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active["B2"].value is None
    assert any("Bronconflict" in w.value for w in at.warning)
    huidig["data"] = Artikeldata(pds, prijslijst=prijzen)
    at.run()
    assert not at.exception and at.get("download_button")
    ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
    assert ws["B2"].value == 340 and ws["C2"].value == 74.58
    assert not any("Bronconflict" in w.value for w in at.warning)
