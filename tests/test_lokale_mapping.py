import asyncio
import io

import openpyxl
import pytest
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from artikeldata import Artikeldata
from dealer_invuller import bepaal_mapping, match_rijen, verwerk, werkboek_naar_bytes


@pytest.fixture
def lokale_data():
    return Artikeldata({"artikelen": {"2010005": {
        "artikelcode": "2010005", "ean": "8714748004368", "omschrijving": "DRY FIX UNI",
        "gn_code": "32141010", "netto_g": 318, "bruto_g": 360,
        "maat_mm": {"l": 89, "b": 48, "h": 184},
        "collo_mm": {"l": 180, "b": 226, "h": 200},
        "min_verkoophoeveelheid": 10,
        "documenten": {"kleur": {"waarde": "transparant", "bron": "productblad"}},
    }}})


def _blad(koppen, rijen=None):
    ws = openpyxl.Workbook().active
    ws.append(koppen)
    for rij in rijen or [["2010005"]]:
        ws.append(rij)
    return ws


@pytest.mark.parametrize("kop,doel,eenheid,waarde", [
    ("Farbton", "kleur", None, "transparant"),
    ("Farbe", "kleur", None, "transparant"),
    ("Kleur", "kleur", None, "transparant"),
    ("Colour", "kleur", None, "transparant"),
    ("Color", "kleur", None, "transparant"),
    ("Artikelbezeichnung / \nArtikelname", "omschrijving", None, "DRY FIX UNI"),
    ("Produktname", "omschrijving", None, "DRY FIX UNI"),
    ("Omschrijving", "omschrijving", None, "DRY FIX UNI"),
    ("Product name", "omschrijving", None, "DRY FIX UNI"),
    ("GN-code", "gn_code", None, "32141010"),
    ("Zolltarifnummer", "gn_code", None, "32141010"),
    ("Commodity code", "gn_code", None, "32141010"),
    ("Nettogewicht (kg)", "netto_gewicht", "kg", 0.318),
    ("Netto gewicht per stuk (gr)", "netto_gewicht", "g", 318),
    ("Net weight [g]", "netto_gewicht", "g", 318),
    ("Bruttogewicht in kg", "bruto_gewicht", "kg", 0.36),
    ("Bruto gewicht per stuk (g)", "bruto_gewicht", "g", 360),
    ("Gross weight (kg)", "bruto_gewicht", "kg", 0.36),
    ("Länge (cm)", "lengte", "cm", 8.9),
    ("Lengte per stuk (mm)", "lengte", "mm", 89),
    ("Length (m)", "lengte", "m", 0.089),
    ("Breite in cm", "breedte", "cm", 4.8),
    ("Breedte (mm)", "breedte", "mm", 48),
    ("Width [mm]", "breedte", "mm", 48),
    ("Höhe (cm)", "hoogte", "cm", 18.4),
    ("Hoogte (mm)", "hoogte", "mm", 184),
    ("Height (m)", "hoogte", "m", 0.184),
    ("Collo lengte (mm)", "collo_lengte", "mm", 180),
    ("Kartonbreite (cm)", "collo_breedte", "cm", 22.6),
    ("Carton height (mm)", "collo_hoogte", "mm", 200),
    ("Minimale verkoophoeveelheid (stuks)", "min_verkoophoeveelheid", "stuks", 10),
])
def test_lokaal_herkende_koppen_vullen_bronwaarde_in(lokale_data, kop, doel, eenheid, waarde):
    ws = _blad(["Artikelcode", kop])
    mapping = bepaal_mapping(None, ws, lokale_data)
    k = mapping.kolommen[1]
    assert (k.doelveld, k.eenheid, k.zekerheid) == (doel, eenheid, "hoog")
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data)
    assert openpyxl.load_workbook(io.BytesIO(inhoud)).active["B2"].value == waarde
    assert rapport.samenvatting()["ingevuld"] == 1


@pytest.mark.parametrize("kop", [
    "Minimale verkoophoeveelheid", "Net weight (cm)", "Hoogte (kg)",
    "Net weight (lb)", "Length (mm/cm)",
])
def test_kwantitatieve_kop_zonder_eenduidige_eenheid_blijft_leeg(lokale_data, kop):
    ws = _blad(["Artikelcode", kop], [["2010005", 0.318], ["2010005", None]])
    mapping = bepaal_mapping(None, ws, lokale_data)
    assert mapping.kolommen[1].doelveld == "geen"
    assert "eenheid" in mapping.kolommen[1].toelichting.lower()


@pytest.mark.parametrize("kop,doel", [
    ("Nettogewicht", "netto_gewicht"), ("Bruttogewicht", "bruto_gewicht"),
    ("Länge", "lengte"), ("Breedte", "breedte"), ("Height", "hoogte"),
    ("Collo lengte", "collo_lengte"),
])
def test_bekende_kop_zonder_eenheid_behoudt_veldherkenning(lokale_data, kop, doel):
    ws = _blad(["Artikelcode", kop])
    mapping = bepaal_mapping(None, ws, lokale_data)
    assert mapping.kolommen[1].doelveld == doel
    assert mapping.kolommen[1].eenheid is None
    assert "geen eenheid" in mapping.kolommen[1].toelichting


@pytest.mark.parametrize("kop", [
    "ArtNr", "Standard-Material", "Artikelnummer", "Farbechtheit", "Farbtonnummer",
    "Gewicht (kg)", "VPE Brutto-Gewicht (kg)",
    "VPE Netto-Gewicht", "Größe Karton", "Spez. Gewicht", "Beschrijving lang",
    "Verwerkingstemperatuur", "Dichtheid", "Verbrauch pro Quadratmeter",
])
def test_ambigue_dealervelden_worden_niet_ingevuld(lokale_data, kop):
    ws = _blad(["Artikelcode", kop])
    mapping = bepaal_mapping(None, ws, lokale_data)
    assert mapping.kolommen[1].doelveld == "geen"


@pytest.fixture
def doos_data():
    return Artikeldata({"artikelen": {"2023205": {
        "artikelcode": "2023205", "omschrijving": "DRY FLEX 1 2-in-1 (150 ml)",
        "netto_g": 172, "min_verkoophoeveelheid": 10,
        "ruw": {"Netto gewicht per doos (kg)": "A: 1,11", "Bruto gewicht per doos (kg)": "2.87",
                "Inhoud (om)doos": "1 x 10"},
        "componenten": [{"naam": "A", "netto_g": 111}, {"naam": "B", "netto_g": 61,
                        "ruw": {"Netto gewicht per doos (kg)": "B: 0,61"}}],
    }}})


VPE_KOPPEN = ["Artikelcode", "VPE Anzahl", "VPE Einheit", "VPE Brutto-Gewicht",
              "VPE Brutto-Gewicht Einheit", "VPE Netto-Gewicht", "VPE Netto-Gewicht Einheit"]


def test_weigel_doosgewicht_wordt_automatisch_numeriek_ingevuld(doos_data):
    ws = _blad(VPE_KOPPEN, [["2023205", 10, "Karton", 2.87, "Kg", None, "Kg"]])
    mapping = bepaal_mapping(None, ws, doos_data)
    assert (mapping.kolommen[3].doelveld, mapping.kolommen[5].doelveld) == (
        "collo_bruto_gewicht", "collo_netto_gewicht",
    )
    assert mapping.kolommen[5].eenheid == "kg"
    uit, rapport = verwerk(werkboek_naar_bytes(ws.parent), "weigel.xlsx", mapping, doos_data,
                          behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(uit))
    assert wb.active["F2"].value == 1.72 and wb.active["D2"].value == 2.87
    assert wb.active["G2"].value == "Kg" and wb.active["F2"].data_type == "n"
    assert rapport.samenvatting()["ingevuld"] == 1
    assert "10" in rapport.rijen[0].velden[1].regel


@pytest.mark.parametrize("kop,doel,verwacht", [
    ("Netto gewicht per doos (kg)", "collo_netto_gewicht", 1.72),
    ("Nettogewicht pro Karton (g)", "collo_netto_gewicht", 1720),
    ("Bruto gewicht per doos (kg)", "collo_bruto_gewicht", 2.87),
    ("Carton gross weight (g)", "collo_bruto_gewicht", 2870),
])
def test_explicitiet_doosgewicht_is_geen_stukgewicht(doos_data, kop, doel, verwacht):
    ws = _blad(["Artikelcode", kop, "Nettogewicht (g)"], [["2023205", None, None]])
    mapping = bepaal_mapping(None, ws, doos_data)
    assert mapping.kolommen[1].doelveld == doel
    uit, _ = verwerk(werkboek_naar_bytes(ws.parent), "doos.xlsx", mapping, doos_data, behoud_sjabloon=True)
    ws = openpyxl.load_workbook(io.BytesIO(uit)).active
    assert ws["B2"].value == verwacht and ws["C2"].value == 172


@pytest.mark.parametrize("index,waarde", [
    (1, 20), (1, None), (1, 0), (1, True), (1, "10-20"),
    (2, "Stk"), (2, "Palette"), (2, None),
    (6, "lb"), (6, None), (6, '=\"Kg\"'),
])
def test_vpe_onjuiste_of_ontbrekende_context_blijft_leeg(doos_data, index, waarde):
    rij = ["2023205", 10, "Karton", 2.87, "Kg", None, "Kg"]
    rij[index] = waarde
    ws = _blad(VPE_KOPPEN, [rij])
    mapping = bepaal_mapping(None, ws, doos_data)
    uit, rapport = verwerk(werkboek_naar_bytes(ws.parent), "doos.xlsx", mapping, doos_data, behoud_sjabloon=True)
    assert openpyxl.load_workbook(io.BytesIO(uit)).active["F2"].value is None
    assert rapport.samenvatting()["onzeker"] == 1


def test_vpe_eenheden_worden_per_rij_gelezen(doos_data):
    from dealer_invuller import ontbrekende_eenheden

    ws = _blad(VPE_KOPPEN, [
        ["2023205", 10, "Karton", 2.87, "Kg", None, eenheid] for eenheid in ("Kg", "g", "lb")
    ])
    mapping = bepaal_mapping(None, ws, doos_data)
    assert mapping.kolommen[5].eenheid is None
    assert not ontbrekende_eenheden(mapping)
    uit, rapport = verwerk(werkboek_naar_bytes(ws.parent), "doos.xlsx", mapping, doos_data, behoud_sjabloon=True)
    ws = openpyxl.load_workbook(io.BytesIO(uit)).active
    assert [ws[f"F{rij}"].value for rij in (2, 3, 4)] == [1.72, 1720, None]
    assert rapport.samenvatting()["ingevuld"] == 2
    assert rapport.samenvatting()["onzeker"] == 1


def test_vpe_gebruikt_minimale_afname_niet_als_doosinhoud(doos_data):
    doos_data.artikelen["2023205"]["ruw"].pop("Inhoud (om)doos")
    ws = _blad(VPE_KOPPEN, [["2023205", 10, "Karton", 2.87, "Kg", None, "Kg"]])
    mapping = bepaal_mapping(None, ws, doos_data)
    uit, rapport = verwerk(werkboek_naar_bytes(ws.parent), "doos.xlsx", mapping, doos_data, behoud_sjabloon=True)
    assert openpyxl.load_workbook(io.BytesIO(uit)).active["F2"].value is None
    assert rapport.samenvatting()["onzeker"] == 1


def test_vpe_dubbele_eenheidkolom_wordt_niet_gegokt(doos_data):
    ws = _blad(VPE_KOPPEN + [VPE_KOPPEN[6]], [["2023205", 10, "Karton", 2.87, "Kg", None, "Kg", "g"]])
    mapping = bepaal_mapping(None, ws, doos_data)
    assert mapping.kolommen[5].doelveld == "geen"


def test_vpe_handmatige_eenheid_tegenstrijdig_met_dealer_blijft_leeg(doos_data):
    ws = _blad(VPE_KOPPEN, [["2023205", 10, "Karton", 2.87, "Kg", None, "Kg"]])
    mapping = bepaal_mapping(None, ws, doos_data)
    mapping.kolommen[5].eenheid = "g"
    uit, rapport = verwerk(werkboek_naar_bytes(ws.parent), "doos.xlsx", mapping, doos_data, behoud_sjabloon=True)
    assert openpyxl.load_workbook(io.BytesIO(uit)).active["F2"].value is None
    assert rapport.samenvatting()["onzeker"] == 1


def test_dubbele_koppen_gebruiken_originele_eenheid(lokale_data):
    ws = _blad(["Artikelcode", "Farbton", "Farbton", "Net weight (kg)", "Net weight (kg)"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    assert [(k.kolom, k.doelveld, k.eenheid) for k in mapping.doelen()] == [
        ("Farbton", "kleur", None), ("Farbton (2)", "kleur", None),
        ("Net weight (kg)", "netto_gewicht", "kg"),
        ("Net weight (kg) (2)", "netto_gewicht", "kg"),
    ]


@pytest.mark.parametrize("marker", ["MW", "MM"])
def test_pim_keuzevelden_blijven_leeg(lokale_data, marker):
    ws = _blad(["Material", marker, "EM"], [
        ["Lief-Mat", "Farbe", "Farbton"], ["2010005", None, None],
    ])
    mapping = bepaal_mapping(None, ws, lokale_data)
    assert mapping.kolommen[1].doelveld == "geen"
    assert mapping.kolommen[2].doelveld == "kleur"


def test_keuzelijst_en_x_velden_zijn_geen_vrije_tekst(lokale_data):
    ws = _blad(["Artikelcode", "Farbe", "Kleur", "Color"], [["2010005", None, "x", None]])
    lijst = DataValidation(type="list", formula1='"rood,groen"')
    ws.add_data_validation(lijst)
    lijst.add(ws["B2"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    assert [k.doelveld for k in mapping.kolommen] == ["sleutel_artikelcode", "geen", "geen", "kleur"]


def test_ean_is_alleen_sleutel_en_omschrijving_nooit_fuzzy_sleutel(lokale_data):
    ws = _blad(["ArtNr", "EAN", "Omschrijving", "Farbton"], [["dealer-1", "8714748004368", "DRY FIX UN", None]])
    mapping = bepaal_mapping(None, ws, lokale_data)
    assert [(k.kolom, k.doelveld) for k in mapping.sleutels()] == [("EAN", "sleutel_ean")]
    assert "EAN" not in [k.kolom for k in mapping.doelen()]
    ws["B2"] = "9999999999999"
    assert match_rijen(ws, mapping, lokale_data)[0].match is None


@pytest.mark.parametrize("fout", [ValueError("API stuk"), asyncio.TimeoutError()])
def test_api_fout_bewaart_lokale_doelvelden(lokale_data, monkeypatch, fout):
    def vraag(*args, **kwargs):
        raise fout

    monkeypatch.setattr("dealer_invuller.vraag_mapping", vraag)
    mapping = bepaal_mapping(object(), _blad(["Artikelcode", "Farbton", "Nettogewicht (kg)"]), lokale_data)
    assert [(k.doelveld, k.eenheid) for k in mapping.doelen()] == [("kleur", None), ("netto_gewicht", "kg")]


@pytest.mark.parametrize("fout", [ValueError("API stuk"), asyncio.TimeoutError()])
def test_geavanceerde_ai_fout_kan_bestaande_mapping_bewaren(lokale_data, monkeypatch, fout):
    def vraag(*args, **kwargs):
        raise fout

    monkeypatch.setattr("dealer_invuller.vraag_mapping", vraag)
    with pytest.raises(type(fout)):
        bepaal_mapping(object(), _blad(["Artikelcode", "Farbton"]), lokale_data, ai_fouten_doorgeven=True)


def test_sjabloonbehoud_bewaart_lege_opmaak_en_bestaand_controleblad(lokale_data):
    ws = _blad(["Artikelcode", "GN-code", "Farbton", "Nettogewicht (kg)"], [
        ["2010005", "bestaand", None, "=1+2"], ["onbekend", None, None, None],
    ])
    ws["C3"].fill = PatternFill("solid", fgColor="008000")
    oude_stijl = ws["C3"]._style
    ws.parent.create_sheet("Controle")["A1"] = "bestaande controle"
    ws.parent.create_sheet("Controle (2)")["A1"] = "ook bewaren"
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(inhoud))
    assert wb.active["B2"].value == "bestaand"
    assert wb.active["C2"].value == "transparant"
    assert wb.active["D2"].value == "=1+2"
    assert wb.active["C3"].value is None and wb.active["C3"]._style == oude_stijl
    assert wb["Controle"]["A1"].value == "bestaande controle"
    assert wb["Controle (2)"]["A1"].value == "ook bewaren"
    assert wb["Controle (3)"]["A1"].value == "Samenvatting"
    assert rapport.samenvatting()["gaten"] == 3


def test_sjabloonbehoud_vult_geen_fuzzy_match_in(lokale_data):
    from mapping import KolomMapping, Mapping

    ws = _blad(["Omschrijving", "Farbton"], [["DRY FIX UN", None]])
    mapping = Mapping(0, [KolomMapping("Omschrijving", "sleutel_omschrijving", None, "laag", ""),
                          KolomMapping("Farbton", "kleur", None, "hoog", "")])
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    assert openpyxl.load_workbook(io.BytesIO(inhoud)).active["B2"].value is None
    assert rapport.rijen[0].velden[0].status == "overgeslagen"
    assert "omschrijving" in rapport.rijen[0].velden[0].regel.lower()


@pytest.mark.parametrize("kleuren", [("groen", "rood"), ("groen", "groen"), ("groen",)])
def test_sjabloonbehoud_laat_componentkleur_onzeker_maar_vult_andere_velden(lokale_data, kleuren):
    artikel = lokale_data.artikelen["2010005"]
    artikel["documenten"] = {}
    artikel["componenten"] = [
        {"naam": naam, "documenten": {"kleur": {"waarde": kleur, "bron": "veiligheidsblad"}}}
        for naam, kleur in zip(("A", "B"), kleuren)
    ]
    ws = _blad(["Artikelcode", "Farbton", "GN-code", "Nettogewicht (kg)"])
    ws["B2"].fill = PatternFill("solid", fgColor="008000")
    oude_stijl = ws["B2"]._style
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(inhoud))
    assert wb.active["B2"].value is None and wb.active["B2"]._style == oude_stijl
    assert wb.active["C2"].value == "32141010"
    assert wb.active["D2"].value == 0.318
    onzeker = rapport.rijen[0].velden[0]
    assert onzeker.status == "onzeker"
    assert "component" in onzeker.bron and "product" in onzeker.regel
    assert rapport.samenvatting()["onzeker"] == 1
    assert rapport.samenvatting()["ingevuld"] == 2


def test_legacy_kan_componentkleur_nog_invullen(lokale_data):
    artikel = lokale_data.artikelen["2010005"]
    artikel["documenten"] = {}
    artikel["componenten"] = [{"naam": "A", "kleur": "groen"}, {"naam": "B", "kleur": "rood"}]
    ws = _blad(["Artikelcode", "Farbton"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data)
    assert openpyxl.load_workbook(io.BytesIO(inhoud)).active["B2"].value == "groen"
    assert rapport.samenvatting()["onzeker"] == 0


def test_productbladkleur_is_wel_eenduidig(lokale_data):
    artikel = lokale_data.artikelen["2010005"]
    artikel["componenten"] = [{"naam": "A", "kleur": "groen"}, {"naam": "B", "kleur": "rood"}]
    ws = _blad(["Artikelcode", "Farbton"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    assert openpyxl.load_workbook(io.BytesIO(inhoud)).active["B2"].value == "transparant"
    assert rapport.samenvatting()["onzeker"] == 0


@pytest.mark.parametrize("kleur", [
    "Standaard wit en reebruin",
    "A oranje transparant, B transparant, gemengd transparante massa",
    "wit of bruin", "weiß oder braun", "white/brown", "rood; groen", "Component A: groen",
])
def test_productblad_met_kleurkeuzes_blijft_onzeker(lokale_data, kleur):
    artikel = lokale_data.artikelen["2010005"]
    artikel["documenten"]["kleur"]["waarde"] = kleur
    ws = _blad(["Artikelcode", "Farbton", "GN-code"])
    ws["B2"].fill = PatternFill("solid", fgColor="008000")
    oude_stijl = ws["B2"]._style
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(inhoud))
    assert wb.active["B2"].value is None and wb.active["B2"]._style == oude_stijl
    assert wb.active["C2"].value == "32141010"
    resultaat = rapport.rijen[0].velden[0]
    assert resultaat.status == "onzeker" and resultaat.bron == "productblad"
    assert "keuzes" in resultaat.regel and "artikelkleur" in resultaat.regel
    assert rapport.samenvatting()["onzeker"] == 1
    # De expliciete handmatige/legacy-route bewaart de onverkorte bronwaarde.
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data)
    assert openpyxl.load_workbook(io.BytesIO(inhoud)).active["B2"].value == kleur


@pytest.mark.parametrize("kop,waarden", [
    ("EAN", ["8714748004368", None]),
    ("EAN", [None, "8714748004368"]),
    ("EAN", ["8714748004368", "87.14748.00436.8"]),
    ("Artikelcode", ["2010005", None]),
    ("Artikelcode", ["2010005", 0]),
    ("Artikelcode", ["2010005", 2010005.0]),
])
def test_dubbele_sleutel_bewaart_gevulde_waarde(lokale_data, kop, waarden):
    ws = _blad([kop, kop, "Farbton"], [waarden])
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    assert openpyxl.load_workbook(io.BytesIO(inhoud)).active["C2"].value == "transparant"
    assert rapport.samenvatting()["gevonden"] == 1


@pytest.mark.parametrize("kop,waarden", [
    ("EAN", ["8714748004368", "8714748003804"]),
    ("EAN", ["8714748003804", "8714748004368"]),
    ("Artikelcode", ["2010005", "2511105"]),
    ("Artikelcode", ["2511105", "2010005"]),
])
def test_tegenstrijdige_dubbele_sleutel_wordt_onzeker_overgeslagen(lokale_data, kop, waarden):
    ws = _blad([kop, kop, "Farbton"], [waarden])
    ws["C2"].fill = PatternFill("solid", fgColor="008000")
    oude_stijl = ws["C2"]._style
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(inhoud))
    assert wb.active["C2"].value is None and wb.active["C2"]._style == oude_stijl
    assert rapport.rijen[0].match is None
    assert rapport.rijen[0].velden[0].status == "onzeker"
    assert rapport.samenvatting()["onzeker"] == 1
    tekst = " ".join(str(c.value) for rij in wb["Controle"] for c in rij if c.value is not None)
    assert "tegenstrijdige" in tekst.casefold()


def test_ontbrekende_eenheden_beperkt_tot_gewichten_en_maten(lokale_data):
    from dealer_invuller import ontbrekende_eenheden
    from mapping import KolomMapping

    ws = _blad(["Artikelcode", "Nettogewicht", "Länge", "Hoogte (cm)", "Farbton", "Gewicht (kg)"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    mapping.kolommen.append(KolomMapping("Afname", "min_verkoophoeveelheid", None, "hoog", ""))
    assert [(k.kolom, k.doelveld) for k in ontbrekende_eenheden(mapping)] == [
        ("Nettogewicht", "netto_gewicht"), ("Länge", "lengte"),
    ]


@pytest.mark.parametrize("maat,gewicht,verwacht", [
    ("mm", "g", [318, 89, 0.36, 18.4]),
    ("cm", "kg", [0.318, 8.9, 0.36, 18.4]),
    ("m", "g", [318, 0.089, 0.36, 18.4]),
])
def test_eenheidskeuze_vult_juist_in_en_behoudt_expliciete_eenheden(lokale_data, maat, gewicht, verwacht):
    from dealer_invuller import ontbrekende_eenheden, pas_eenheden_toe

    lokale_data.artikelen["2010005"]["netto_regel"] = "222 + 96 = 318 g"
    ws = _blad(["Artikelcode", "Nettogewicht", "Länge", "Bruttogewicht (kg)", "Höhe (cm)"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    origineel = mapping.naar_dict()
    gekozen = pas_eenheden_toe(mapping, maat_eenheid=maat, gewicht_eenheid=gewicht,
                              bron="Profiel Primärlieferant973184")
    assert mapping.naar_dict() == origineel
    assert gekozen is not mapping and gekozen.kolommen[0] is not mapping.kolommen[0]
    assert ontbrekende_eenheden(gekozen) == []
    assert gekozen.kolommen[3].eenheid == "kg" and gekozen.kolommen[4].eenheid == "cm"
    assert gekozen.kolommen[3].toelichting == mapping.kolommen[3].toelichting
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", gekozen, lokale_data,
                             behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(inhoud))
    assert [wb.active.cell(2, kolom).value for kolom in range(2, 6)] == verwacht
    assert rapport.samenvatting()["ingevuld"] == 4 and rapport.samenvatting()["eenheid_nodig"] == 0
    gewicht_resultaat = rapport.rijen[0].velden[0]
    assert "222 + 96 = 318 g" in gewicht_resultaat.regel
    assert "Profiel Primärlieferant973184" in gewicht_resultaat.regel
    assert gekozen.kolommen[1].toelichting in gewicht_resultaat.regel
    assert "Profiel Primärlieferant973184" in " ".join(str(c.value) for rij in wb["Controle"] for c in rij)


def test_eenheidskeuze_vult_alleen_aangegeven_dimensie_in(lokale_data):
    from dealer_invuller import ontbrekende_eenheden, pas_eenheden_toe

    ws = _blad(["Artikelcode", "Nettogewicht", "Länge"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    gekozen = pas_eenheden_toe(mapping, gewicht_eenheid="g")
    assert [k.kolom for k in ontbrekende_eenheden(gekozen)] == ["Länge"]
    assert "Keuze gebruiker" in gekozen.kolommen[1].toelichting
    assert mapping.kolommen[1].eenheid is None


@pytest.mark.parametrize("keuze", [
    {"maat_eenheid": "g"}, {"maat_eenheid": "inch"}, {"maat_eenheid": ""},
    {"gewicht_eenheid": "mm"}, {"gewicht_eenheid": "lb"}, {"gewicht_eenheid": ""},
])
def test_ongeldige_eenheidskeuze_geeft_fout_zonder_mutatie(lokale_data, keuze):
    from dealer_invuller import pas_eenheden_toe

    mapping = bepaal_mapping(None, _blad(["Artikelcode", "Nettogewicht", "Länge"]), lokale_data)
    origineel = mapping.naar_dict()
    with pytest.raises(ValueError, match="eenheid"):
        pas_eenheden_toe(mapping, **keuze)
    assert mapping.naar_dict() == origineel


def test_ontbrekende_eenheid_slaat_alleen_die_cellen_over_en_bewaart_opmaak(lokale_data):
    ws = _blad(["Artikelcode", "Nettogewicht", "Länge", "GN-code"], [
        ["2010005", None, None, None], ["2010005", 999, "=1+2", None],
    ])
    ws["B2"].fill = PatternFill("solid", fgColor="008000")
    ws["C2"].number_format = "0.000"
    stijlen = [ws["B2"]._style, ws["C2"]._style]
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data,
                             behoud_sjabloon=True)
    wb = openpyxl.load_workbook(io.BytesIO(inhoud))
    assert [wb.active["B2"].value, wb.active["C2"].value] == [None, None]
    assert [wb.active["B2"]._style, wb.active["C2"]._style] == stijlen
    assert wb.active["D2"].value == "32141010" and wb.active["D3"].value == "32141010"
    assert wb.active["B3"].value == 999 and wb.active["C3"].value == "=1+2"
    assert [v.status for v in rapport.rijen[0].velden] == ["eenheid_nodig", "eenheid_nodig", "ingevuld"]
    assert rapport.samenvatting()["eenheid_nodig"] == 2
    assert rapport.samenvatting()["ingevuld"] == 2


def test_legacy_zonder_eenheid_bewaart_bestaande_broneenheid_invulling(lokale_data):
    ws = _blad(["Artikelcode", "Nettogewicht", "Länge"])
    mapping = bepaal_mapping(None, ws, lokale_data)
    inhoud, rapport = verwerk(werkboek_naar_bytes(ws.parent), "dealer.xlsx", mapping, lokale_data)
    wb = openpyxl.load_workbook(io.BytesIO(inhoud))
    assert wb.active["B2"].value == 318 and wb.active["C2"].value == 89
    assert rapport.samenvatting()["eenheid_nodig"] == 0
