"""Aanvullende adviesprijzen mogen technische data niet veranderen of inkoopprijzen vullen."""

import io
from datetime import date
from types import SimpleNamespace

import openpyxl
import pytest
from streamlit.testing.v1 import AppTest

import app
import artikeldata
from artikeldata import Artikeldata
from dealer_invuller import bepaal_mapping, verwerk, werkboek_naar_bytes


@pytest.fixture
def prijsdata(monkeypatch):
    class Vandaag(date):
        @classmethod
        def today(cls):
            return cls(2026, 9, 7)

    monkeypatch.setattr(artikeldata, "date", Vandaag)
    monkeypatch.setattr(app, "date", Vandaag, raising=False)
    prijzen = {
        "bron": "prijzen_2026.csv", "geldig_vanaf": "2026-01-01", "geldig_tot": "2026-12-31",
        "valuta": "EUR", "btw": "exclusief", "meldingen": [],
        "artikelen": {
            "2023205": {"omschrijving": "DRY FLEX 1 2-in-1", "ean": "8714748004917",
                        "ve_aantal": 10, "eenheid": "st", "adviesprijs_cent": 4509, "bronregel": 7},
            "2040005": {"omschrijving": "Universele Kleurpigmenten", "ean": "8714748005112",
                        "ve_aantal": 4, "eenheid": "set", "adviesprijs_cent": 2299, "bronregel": 22},
            "2023999": {"omschrijving": "DRY FLEX 1", "ean": "8714748004740",
                        "ve_aantal": 20, "eenheid": "set", "adviesprijs_cent": 5000, "bronregel": 6},
        },
    }
    return Artikeldata({"artikelen": {
        "2023205": {"artikelcode": "2023205", "omschrijving": "Technische productnaam",
                    "ean": "8714748004917", "netto_g": 172},
        "2023999": {"artikelcode": "2023999", "omschrijving": "DRY FLEX 1",
                    "ean": "8714748002616", "netto_g": 340},
    }}, prijslijst=prijzen)


def _invullen(prijsdata, koppen, rijen):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(koppen)
    for rij in rijen:
        ws.append(rij)
    mapping = bepaal_mapping(None, ws, prijsdata)
    uit, rapport = verwerk(werkboek_naar_bytes(wb), "dealer.xlsx", mapping, prijsdata, behoud_sjabloon=True)
    return openpyxl.load_workbook(io.BytesIO(uit)), rapport, mapping


@pytest.mark.parametrize("kop", [
    "Adviesprijs excl. btw (EUR)", "Verkoopadviesprijs excl. btw (€)",
    "UVP netto [EUR]", "Recommended retail price excl. VAT (EUR)",
])
def test_explicitiete_adviesprijs_is_numeriek_met_bron_en_prijseenheid(prijsdata, kop):
    wb, rapport, mapping = _invullen(prijsdata, ["Artikelcode", kop], [["2040005"], ["2023205"]])
    assert mapping.kolommen[1].doelveld == "adviesprijs"
    assert mapping.kolommen[1].eenheid == "EUR"
    assert [wb.active[f"B{r}"].value for r in (2, 3)] == [22.99, 45.09]
    assert wb.active["B2"].data_type == "n"
    assert "prijzen_2026.csv" in rapport.rijen[0].velden[0].bron
    assert "exclusief btw" in rapport.rijen[0].velden[0].regel
    assert "per set" in rapport.rijen[0].velden[0].regel


@pytest.mark.parametrize("kop", [
    "Prijs (EUR)", "Inkoopprijs (EUR)", "Dealerprijs excl. btw (EUR)", "UVP (EUR)",
    "Adviesprijs incl. btw (EUR)", "Adviesprijs excl. btw (GBP)", "Adviesprijs excl. btw",
    "Adviesprijs excl. btw per stuk (EUR)", "Adviesprijs excl. btw ($)",
    "VE", "VPE Anzahl", "Doosinhoud",
])
def test_ambigue_prijs_of_verpakking_blijft_leeg(prijsdata, kop):
    wb, _, mapping = _invullen(prijsdata, ["Artikelcode", kop], [["2040005"]])
    assert mapping.kolommen[1].doelveld == "geen"
    assert wb.active["B2"].value is None


def test_nieuw_artikel_krijgt_alleen_werkelijke_brongegevens(prijsdata):
    wb, rapport, _ = _invullen(prijsdata, [
        "EAN", "Omschrijving", "Nettogewicht (g)", "Lengte (mm)",
        "Adviesprijs excl. btw (EUR)", "Adviesprijs eenheid", "VE volgens prijslijst",
    ], [["8714748005112"]])
    assert rapport.rijen[0].match.via == "ean"
    assert [wb.active.cell(2, c).value for c in range(2, 8)] == [
        "Universele Kleurpigmenten", None, None, 22.99, "set", 4,
    ]


@pytest.mark.parametrize("kop", ["Omschrijving", "Productnaam", "Produktname", "ArtBeschreibung"])
def test_alle_naamkolommen_gebruiken_prijslijst_met_behoud_van_dealertekst(prijsdata, kop):
    wb, rapport, _ = _invullen(prijsdata, ["Artikelcode", kop, "Nettogewicht (g)"], [
        ["2023205"], ["2023205", "Eigen dealeromschrijving"], ["2040005"],
    ])
    assert [wb.active[f"B{r}"].value for r in (2, 3, 4)] == [
        "DRY FLEX 1 2-in-1", "Eigen dealeromschrijving", "Universele Kleurpigmenten",
    ]
    assert wb.active["C2"].value == 172
    assert rapport.rijen[0].velden[0].bron == "prijzen_2026.csv, rij 7"
    assert "leidend" in rapport.rijen[0].velden[0].regel


def test_bewust_overschrijven_gebruikt_prijslijstomschrijving(prijsdata):
    wb = openpyxl.Workbook()
    wb.active.append(["Artikelcode", "Omschrijving"])
    wb.active.append(["2023205", "Oude omschrijving"])
    mapping = bepaal_mapping(None, wb.active, prijsdata)
    uit, _ = verwerk(werkboek_naar_bytes(wb), "dealer.xlsx", mapping, prijsdata,
                     overschrijven=True, behoud_sjabloon=True)
    assert openpyxl.load_workbook(io.BytesIO(uit)).active["B2"].value == "DRY FLEX 1 2-in-1"


def test_bronconflict_blokkeert_rij_en_behoudt_bestaande_waarden(prijsdata):
    wb, rapport, _ = _invullen(prijsdata, [
        "Artikelcode", "Adviesprijs excl. btw (EUR)", "Nettogewicht (g)", "Omschrijving",
    ], [["2023999", None, None, "Naam dealer"]])
    assert [wb.active.cell(2, c).value for c in (2, 3, 4)] == [None, None, "Naam dealer"]
    assert "8714748002616" in rapport.rijen[0].toelichting
    controle = " ".join(str(c.value) for rij in wb["Controle"] for c in rij if c.value is not None)
    assert "8714748004740" in controle
    assert rapport.samenvatting()["onzeker"] == 2


def _start_app():
    import app
    app.main()


@pytest.fixture
def prijzen_app(monkeypatch, tmp_path, prijsdata):
    monkeypatch.setattr(app, "get_secret", lambda naam: None)
    monkeypatch.setattr(app.dealer_profielen, "PROFIEL_MAP", tmp_path / "profielen")
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: prijsdata)
    bestand = SimpleNamespace(name="prijzen.csv", getvalue=lambda: (
        "Artikelcode;Adviesprijs excl. btw (EUR);Nettogewicht (g)\n2040005;;\n2023205;;\n2023999;;\n"
    ).encode())
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)

    def geen_ai(*args, **kwargs):
        raise AssertionError("Deze invulstap mag geen AI gebruiken.")

    monkeypatch.setattr(app.anthropic, "AsyncAnthropic", geen_ai)
    return AppTest.from_function(_start_app)


def test_prijslijst_werkt_direct_en_bronconflict_is_zichtbaar(prijzen_app):
    at = prijzen_app.run()
    assert not at.exception and at.get("download_button")
    ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
    assert [ws[f"B{r}"].value for r in (2, 3, 4)] == [22.99, 45.09, None]
    assert ws["C3"].value == 172 and ws["C4"].value is None
    assert not any("prijzen_2026.csv" in c.value for c in at.caption)
    assert any("2023999" in w.value and "bron" in w.value.lower() for w in at.warning)
    assert not any(e.label.startswith("Broncontrole") for e in at.expander)


def test_download_vernieuwt_na_geldigheidsdatum(prijzen_app, monkeypatch):
    at = prijzen_app.run()
    assert not at.exception
    assert openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active["B2"].value == 22.99
    class NaGeldigheid(date):
        @classmethod
        def today(cls):
            return cls(2027, 1, 1)
    monkeypatch.setattr(artikeldata, "date", NaGeldigheid)
    monkeypatch.setattr(app, "date", NaGeldigheid, raising=False)
    at.run()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
    assert ws["B2"].value is None and ws["B3"].value is None
    assert ws["C3"].value == 172


def test_gewijzigde_bronprijs_vernieuwt_download(prijzen_app, prijsdata):
    at = prijzen_app.run()
    assert not at.exception
    prijsdata.artikelen["2040005"]["prijslijst"]["adviesprijs_cent"] = 2499
    at.run()
    assert not at.exception
    assert openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active["B2"].value == 24.99


@pytest.mark.parametrize("handmatig_overslaan", [False, True])
def test_oude_prijskoppeling_vernieuwt_met_behoud_handmatige_keuzes(prijzen_app, monkeypatch, handmatig_overslaan):
    import json

    oorspronkelijke_mapping = app.bepaal_mapping

    def oude_mapping(*args, **kwargs):
        mapping = oorspronkelijke_mapping(*args, **kwargs)
        for kolom in mapping.kolommen:
            if kolom.doelveld == "adviesprijs":
                kolom.doelveld, kolom.eenheid = "geen", None
        return mapping

    monkeypatch.setattr(app, "bepaal_mapping", oude_mapping)
    at = prijzen_app.run()
    assert not at.exception
    assert openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active["B2"].value is None
    # De browser bewaart wijzigingen aan de tabel apart van de oorspronkelijke mapping.
    rijen = {"2": {"Doelveld": "Niet invullen  [geen]"}}
    if handmatig_overslaan:
        rijen["1"] = {"Doelveld": "Niet invullen  [geen]"}
    wijzigingen = {"edited_rows": rijen, "added_rows": [], "deleted_rows": []}

    def herhaal_met_keuzes():
        widgets = at._tree.get_widget_states()
        widgets.widgets.add(id=at.dataframe[0].proto.id, string_value=json.dumps(wijzigingen))
        at._run(widgets)

    herhaal_met_keuzes()
    assert not at.exception
    at.session_state["dealer"]["mapping_versie"] = "doosgewichten_v1"
    monkeypatch.setattr(app, "bepaal_mapping", oorspronkelijke_mapping)
    herhaal_met_keuzes()
    assert not at.exception
    ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
    assert ws["B2"].value == (None if handmatig_overslaan else 22.99)
    assert ws["C3"].value is None


def test_onleesbare_prijslijst_geeft_melding_en_technische_data_blijft_werken(prijzen_app, prijsdata):
    prijsdata.prijslijst_info = {}
    prijsdata.bron_meldingen = ["De aanvullende prijslijst kon niet worden gelezen."]
    for artikel in prijsdata.artikelen.values():
        artikel.pop("prijslijst", None)
        artikel.pop("bron_conflicten", None)
    at = prijzen_app.run()
    assert not at.exception
    assert any("Brongegevens konden niet volledig worden geladen" in w.value for w in at.warning)
    ws = openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"])).active
    assert ws["B3"].value is None and ws["C3"].value == 172
