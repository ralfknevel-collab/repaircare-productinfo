"""Duitse productteksten in de echte upload- en downloadstroom."""

import io
import json
from types import SimpleNamespace

import openpyxl
import pytest
from streamlit.testing.v1 import AppTest

import app
from artikeldata import Artikeldata


@pytest.fixture
def producttekst_app(monkeypatch, tmp_path):
    for naam in ("APP_LANGUAGE", "RENDER", "REQUIRE_APP_PASSWORD"):
        monkeypatch.delenv(naam, raising=False)
    monkeypatch.setattr(app, "get_secret", lambda naam: None)
    monkeypatch.setattr(app.dealer_profielen, "PROFIEL_MAP", tmp_path / "profielen")
    data = Artikeldata({"artikelen": {"2511105": {
        "artikelcode": "2511105", "omschrijving": "DRY SEAL™ MP Wit",
        "ean": "8714748003804", "netto_g": 452, "gn_code": "32141010",
    }}})
    monkeypatch.setattr(app.Artikeldata, "laad", lambda: data)
    bestand = SimpleNamespace(name="dealer.csv", getvalue=lambda: (
        "Artikelcode;Omschrijving;Nettogewicht (g);Extra\n2511105;;;\n"
    ).encode("utf-8"))
    monkeypatch.setattr(app.st, "file_uploader", lambda *args, **kwargs: bestand)

    def geen_netwerk(*args, **kwargs):
        raise AssertionError("Voor productteksten mag geen AI-client worden gestart.")

    monkeypatch.setattr(app.anthropic, "Anthropic", geen_netwerk)
    monkeypatch.setattr(app.anthropic, "AsyncAnthropic", geen_netwerk)
    return AppTest.from_string("import app\napp.main()"), data, bestand


def werkboek(at):
    return openpyxl.load_workbook(io.BytesIO(at.session_state["dealer"]["uit"]))


def test_duitse_upload_vult_vertaalde_omschrijving_met_ongewijzigd_gewicht(producttekst_app, monkeypatch):
    # Vangt een ontbrekende koppeling tussen de taalkeuze en de invulstap.
    at, data, _ = producttekst_app
    monkeypatch.setenv("APP_LANGUAGE", "de")
    at.run()
    assert not at.exception
    wb = werkboek(at)
    assert wb.active["B2"].value == "DRY SEAL™ MP Weiß"
    assert wb.active["C2"].value == 452
    assert data.artikelen["2511105"]["omschrijving"] == "DRY SEAL™ MP Wit"
    controle = " ".join(str(c.value) for rij in wb["Controle"] for c in rij)
    assert "DRY SEAL™ MP Weiß" in controle
    assert "vertal" in controle.lower()


def test_taalwissel_vernieuwt_tekst_en_behoudt_handmatige_koppeling(producttekst_app):
    # Vangt hergebruik van de Nederlandse download of verlies van de mapping.
    at, _, _ = producttekst_app
    at.run()
    assert werkboek(at).active["B2"].value == "DRY SEAL™ MP Wit"
    edits = {"edited_rows": {"3": {"Doelveld": "Douanetariefnummer (GN/HS-code)  [gn_code]"}},
             "added_rows": [], "deleted_rows": []}
    widgets = at._tree.get_widget_states()
    widgets.widgets.add(id=at.dataframe[0].proto.id, string_value=json.dumps(edits))
    at._run(widgets)
    for taal, tekst in (("de", "DRY SEAL™ MP Weiß"), ("nl", "DRY SEAL™ MP Wit")):
        at.radio[0].set_value(taal).run()
        assert not at.exception
        assert werkboek(at).active["B2"].value == tekst
        assert werkboek(at).active["D2"].value == "32141010"


def test_gewijzigde_omschrijving_meldt_ontbrekende_vertaling_en_vult_gewicht(producttekst_app, monkeypatch):
    # Vangt de oude vertaling op alleen artikelcode of een stil Nederlands resultaat.
    at, data, _ = producttekst_app
    monkeypatch.setenv("APP_LANGUAGE", "de")
    at.run()
    data.artikelen["2511105"]["omschrijving"] = "Nieuwe Nederlandse productnaam"
    at.run()
    assert not at.exception
    wb = werkboek(at)
    assert wb.active["B2"].value is None
    assert wb.active["C2"].value == 452
    assert any("Übersetzung" in bericht.value for bericht in at.warning)
    assert at.session_state["dealer"]["rapport"].samenvatting()["vertaling_ontbreekt"] == 1


def test_bestaande_dealertekst_wordt_niet_vertaald(producttekst_app, monkeypatch):
    # Vangt een vertaalronde over alle cellen in plaats van alleen nieuwe invulwaarden.
    at, _, bestand = producttekst_app
    bestand.getvalue = lambda: "Artikelcode;Omschrijving;Nettogewicht (g)\n2511105;Eigen dealertekst;\n".encode()
    monkeypatch.setenv("APP_LANGUAGE", "de")
    at.run()
    assert not at.exception
    assert werkboek(at).active["B2"].value == "Eigen dealertekst"
    assert werkboek(at).active["C2"].value == 452


def test_gewijzigde_catalogus_vernieuwt_download_en_laadfout_verwijdert_oude_tekst(producttekst_app, monkeypatch):
    # Vangt caching op alleen de versie en behoud van oude tekst na een laadfout.
    from productteksten import Productteksten
    at, _, _ = producttekst_app
    monkeypatch.setenv("APP_LANGUAGE", "de")
    at.run()
    assert werkboek(at).active["B2"].value == "DRY SEAL™ MP Weiß"
    vervanging = Productteksten({"omschrijving": {"DRY SEAL™ MP Wit": "DRY SEAL™ MP weiß"}})
    monkeypatch.setattr(app.Productteksten, "laad", lambda: vervanging)
    at.run()
    assert werkboek(at).active["B2"].value == "DRY SEAL™ MP weiß"

    def kapotte_catalogus():
        raise ValueError("Ongeldige catalogus")

    monkeypatch.setattr(app.Productteksten, "laad", kapotte_catalogus)
    at.run()
    assert not at.exception
    assert werkboek(at).active["B2"].value is None
    assert werkboek(at).active["C2"].value == 452
    assert any("nicht geladen" in bericht.value for bericht in at.warning)


@pytest.mark.parametrize("overschrijven", [False, True])
def test_formules_en_bronvoorrang_blijven_intact_bij_vertalen(overschrijven):
    # Vangt vertaling vóór de formulecheck of vóór de leidende prijslijstkeuze.
    from dealer_invuller import bepaal_mapping, verwerk, werkboek_naar_bytes
    prijzen = {"bron": "prijs.csv", "meldingen": [], "geldig_vanaf": "2026-01-01",
               "geldig_tot": "2026-12-31", "valuta": "EUR", "btw": "exclusief", "artikelen": {
        "2511105": {"omschrijving": "DRY SEAL™ MP Wit", "ean": "8714748003804", "bronregel": 7,
                    "adviesprijs_cent": 2299, "ve_aantal": 12, "eenheid": "st"},
    }}
    data = Artikeldata({"artikelen": {"2511105": {
        "artikelcode": "2511105", "omschrijving": "Oude productnaam", "ean": "8714748003804", "netto_g": 452,
    }}}, prijslijst=prijzen)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Artikelcode", "Omschrijving", "Produktname", "Nettogewicht (g)"])
    ws.append(["2511105", '="Dealerformule"', None, None])
    origineel = werkboek_naar_bytes(wb)
    mapping = bepaal_mapping(None, ws, data)
    uit, rapport = verwerk(origineel, "dealer.xlsx", mapping, data,
                          overschrijven=overschrijven, behoud_sjabloon=True, producttaal="de")
    ingevuld = openpyxl.load_workbook(io.BytesIO(uit)).active
    assert ingevuld["B2"].data_type == "f" and ingevuld["B2"].value == '="Dealerformule"'
    assert ingevuld["C2"].value == "DRY SEAL™ MP Weiß"
    assert ingevuld["D2"].value == 452
    naamveld = next(v for v in rapport.rijen[0].velden if v.kolom == "Produktname")
    assert "prijs.csv" in naamveld.bron and "vertal" in naamveld.regel.lower()


def test_ontbrekende_vertaling_wist_geen_bestaande_waarde_bij_overschrijven(producttekst_app, monkeypatch):
    # Vangt wissen van bestaande inhoud wanneer de nieuwe brontekst niet vertaald is.
    at, data, bestand = producttekst_app
    data.artikelen["2511105"]["omschrijving"] = "Nieuw product zonder vertaling"
    bestand.getvalue = lambda: "Artikelcode;Omschrijving;Nettogewicht (g)\n2511105;Bestaande productnaam;\n".encode()
    monkeypatch.setenv("APP_LANGUAGE", "de")
    at.run()
    next(c for c in at.checkbox if "überschreiben" in c.label).set_value(True).run()
    assert not at.exception
    assert werkboek(at).active["B2"].value == "Bestaande productnaam"
    assert werkboek(at).active["C2"].value == 452
    assert any("Übersetzung" in bericht.value for bericht in at.warning)
